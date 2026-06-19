import io
import csv
import hashlib
import unicodedata
from decimal import Decimal, InvalidOperation
import json
from collections import defaultdict

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from core.models import HomologacaoSetorItem, Pregao, PregaoItem, Setor
from core.rubricas import rubrica_normalizada
from core.views import gestor_financeiro_required

from .models import (
    RecursoOrcamentario, Transferencia, Despesa,
    PdiPerspectiva, PdiObjetivoEstrategico, PdiIndicador,
    NaturezaRecurso, Rubrica, OrigemRecurso, SituacaoDespesa,
    RegistroPrecoVigente, RegistroPrecoItem,
)
from .forms import (
    RecursoOrcamentarioForm, TransferenciaForm, DespesaForm,
    PdiPerspectivaForm, PdiObjetivoForm, PdiIndicadorForm,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _anos_disponiveis():
    """Retorna anos de 2020 ao corrente, mais quaisquer anos já cadastrados fora do range."""
    from datetime import date
    ano_atual = date.today().year
    fixos = set(range(2020, ano_atual + 1))
    existentes = set(
        RecursoOrcamentario.objects.values_list('ano_fiscal', flat=True).distinct()
    )
    return sorted(fixos | existentes)


def _setor_catalog_label(setor):
    if not setor:
        return ''
    base = f'{setor.sigla} — {setor.nome}' if setor.sigla else setor.nome
    return f'{base} ({setor.codigo})' if setor.codigo else base


def _filtros_dashboard(request):
    return {
        'ano':      request.GET.getlist('ano'),
        'natureza': request.GET.get('natureza', ''),
        'rubrica':  request.GET.get('rubrica', ''),
        'origem':   request.GET.get('origem', ''),
        'setor':    request.GET.get('setor', ''),
    }


def _aplicar_filtros_recursos(qs, filtros):
    if filtros['ano']:
        qs = qs.filter(ano_fiscal__in=filtros['ano'])
    if filtros['natureza']:
        qs = qs.filter(natureza__iexact=filtros['natureza'])
    if filtros['rubrica']:
        qs = qs.filter(rubrica=filtros['rubrica'])
    if filtros['origem']:
        qs = qs.filter(origem_recurso=filtros['origem'])
    if filtros['setor']:
        qs = qs.filter(setor__pk=filtros['setor'])
    return qs


def _aplicar_filtros_despesa(qs, filtros):
    from django.db.models import Q
    if filtros['ano']:
        qs = qs.filter(data_despesa__year__in=filtros['ano'])
    if filtros['natureza']:
        qs = qs.filter(natureza__iexact=filtros['natureza'])
    if filtros['rubrica']:
        qs = qs.filter(rubrica=filtros['rubrica'])
    if filtros['origem']:
        qs = qs.filter(recurso__origem_recurso=filtros['origem'])
    if filtros['setor']:
        qs = qs.filter(
            Q(setor__pk=filtros['setor']) | Q(recurso__setor__pk=filtros['setor'])
        )
    return qs


def _dashboard_natureza_key(natureza):
    valor = str(natureza or '').strip().lower()
    if 'capital' in valor:
        return 'capital'
    if 'custeio' in valor:
        return 'custeio'
    return valor


def _dashboard_setor_label(setor):
    if not setor:
        return 'Sem setor'
    return setor.sigla or setor.nome or 'Sem setor'


def _saldos_recursos(recursos):
    ids = [r.pk for r in recursos]
    zero = Decimal('0')
    if not ids:
        return {}

    entradas = {
        row['destino_id']: row['total'] or zero
        for row in Transferencia.objects.filter(status='realizada', destino_id__in=ids)
        .values('destino_id')
        .annotate(total=Sum('valor'))
    }
    saidas = {
        row['origem_id']: row['total'] or zero
        for row in Transferencia.objects.filter(status='realizada', origem_id__in=ids)
        .values('origem_id')
        .annotate(total=Sum('valor'))
    }
    situacoes_sem_impacto = SituacaoDespesa.chaves_sem_impacto()
    despesas = {
        row['recurso_id']: row['total'] or zero
        for row in Despesa.objects.exclude(situacao__in=situacoes_sem_impacto)
        .filter(recurso_id__in=ids)
        .values('recurso_id')
        .annotate(total=Sum('valor_comprometido'))
    }

    saldos = {}
    for recurso in recursos:
        saldos[recurso.pk] = (
            (recurso.valor_orcamentario or zero)
            + entradas.get(recurso.pk, zero)
            - saidas.get(recurso.pk, zero)
            - despesas.get(recurso.pk, zero)
        )
    return saldos


# ── Dashboard público ─────────────────────────────────────────────────────────

def orcamento_dashboard(request):
    filtros = _filtros_dashboard(request)
    anos_selecionados = filtros['ano']

    recursos_qs = _aplicar_filtros_recursos(
        RecursoOrcamentario.objects.select_related('setor'),
        filtros
    )
    despesas_qs = _aplicar_filtros_despesa(
        Despesa.objects.exclude(situacao__in=SituacaoDespesa.chaves_sem_impacto()).select_related('setor', 'recurso'),
        filtros
    )

    recursos_lista = list(recursos_qs)
    saldos_recursos = _saldos_recursos(recursos_lista)

    total_orcado_bruto = recursos_qs.aggregate(t=Sum('valor_orcamentario'))['t'] or Decimal('0')
    # Ajuste por transferências: entradas somam, saídas subtraem (só nos recursos filtrados)
    rec_ids = [r.pk for r in recursos_lista]
    transf_entradas = (Transferencia.objects.filter(status='realizada', destino_id__in=rec_ids)
                       .aggregate(t=Sum('valor'))['t'] or Decimal('0'))
    transf_saidas = (Transferencia.objects.filter(status='realizada', origem_id__in=rec_ids)
                     .aggregate(t=Sum('valor'))['t'] or Decimal('0'))
    total_orcado = total_orcado_bruto + transf_entradas - transf_saidas
    total_comprometido = despesas_qs.aggregate(t=Sum('valor_comprometido'))['t'] or Decimal('0')
    saldo = sum(saldos_recursos.values(), Decimal('0'))
    pct_execucao = (float(total_comprometido) / float(total_orcado) * 100) if total_orcado else 0

    def _fmt_brl(v):
        """Formata float como R$ 1.234.567,89 (pt-BR)."""
        import locale as _locale
        # Formata manualmente para garantir pt-BR independente da config do servidor
        v = float(v)
        neg = v < 0
        s = f'{abs(v):,.2f}'          # 1,234,567.89
        s = s.replace(',', 'X').replace('.', ',').replace('X', '.')  # 1.234.567,89
        return ('-' if neg else '') + 'R$ ' + s

    kpi = {
        'total_orcado': float(total_orcado),
        'total_comprometido': float(total_comprometido),
        'saldo': float(saldo),
        'pct_execucao': round(pct_execucao, 1),
        'total_orcado_fmt': _fmt_brl(total_orcado),
        'total_comprometido_fmt': _fmt_brl(total_comprometido),
        'saldo_fmt': _fmt_brl(saldo),
    }

    # — Gráfico 1: Natureza por ano —
    anos_area = sorted({r.ano_fiscal for r in recursos_lista})
    natureza_map = {
        'custeio': defaultdict(Decimal),
        'capital': defaultdict(Decimal),
    }
    for recurso in recursos_lista:
        chave = _dashboard_natureza_key(recurso.natureza)
        if chave in natureza_map:
            natureza_map[chave][recurso.ano_fiscal] += saldos_recursos.get(recurso.pk, Decimal('0'))

    dados_area = json.dumps({
        'anos':    anos_area,
        'custeio': [float(natureza_map['custeio'].get(a, Decimal('0'))) for a in anos_area],
        'capital': [float(natureza_map['capital'].get(a, Decimal('0'))) for a in anos_area],
    })

    # — Gráfico 2: Setor —
    rec_setor = defaultdict(Decimal)
    for recurso in recursos_lista:
        rec_setor[_dashboard_setor_label(recurso.setor)] += saldos_recursos.get(recurso.pk, Decimal('0'))

    comp_setor = defaultdict(Decimal)
    for row in despesas_qs.values('setor__sigla', 'setor__nome').annotate(total=Sum('valor_comprometido')):
        label = row['setor__sigla'] or row['setor__nome'] or 'Sem setor'
        comp_setor[label] += row['total'] or Decimal('0')

    setor_labels = sorted(
        set(rec_setor) | set(comp_setor),
        key=lambda label: rec_setor.get(label, Decimal('0')) + comp_setor.get(label, Decimal('0')),
        reverse=True,
    )[:15]
    # Calcular disponível por setor
    dados_setor = json.dumps({
        'labels':      setor_labels,
        'comprometido':[float(comp_setor.get(label, Decimal('0'))) for label in setor_labels],
        'disponivel':  [float(rec_setor.get(label, Decimal('0'))) for label in setor_labels],
    })

    # — Gráfico 3: Rubrica —
    rec_rubrica = defaultdict(Decimal)
    for recurso in recursos_lista:
        rec_rubrica[recurso.rubrica or 'Sem rubrica'] += saldos_recursos.get(recurso.pk, Decimal('0'))

    comp_rubrica = defaultdict(Decimal)
    for row in despesas_qs.values('rubrica').annotate(total=Sum('valor_comprometido')):
        label = row['rubrica'] or 'Sem rubrica'
        comp_rubrica[label] += row['total'] or Decimal('0')

    rubrica_labels = sorted(
        set(rec_rubrica) | set(comp_rubrica),
        key=lambda label: rec_rubrica.get(label, Decimal('0')) + comp_rubrica.get(label, Decimal('0')),
        reverse=True,
    )
    dados_rubrica = json.dumps({
        'labels':      rubrica_labels,
        'comprometido':[float(comp_rubrica.get(label, Decimal('0'))) for label in rubrica_labels],
        'disponivel':  [float(rec_rubrica.get(label, Decimal('0'))) for label in rubrica_labels],
    })

    # — Gráfico PDI: execução por Perspectiva / Objetivo Estratégico —
    pdi_perspectiva_map = {}
    sem_vinculo_pdi = Decimal('0')

    # pdi_vinculos column may not exist yet (pending migration); load it safely
    pdi_vinculos_map = {}
    try:
        for row in despesas_qs.values('id', 'pdi_vinculos'):
            pdi_vinculos_map[row['id']] = row['pdi_vinculos'] or []
    except Exception:
        pass

    for despesa in despesas_qs.defer('pdi_vinculos').select_related('perspectiva_pdi', 'objetivo_pdi'):
        valor = despesa.valor_comprometido or Decimal('0')
        vinculos = pdi_vinculos_map.get(despesa.id) or []

        if vinculos:
            share = valor / Decimal(len(vinculos))
            for link in vinculos:
                p_id = link.get('perspectiva_id')
                p_nome = link.get('perspectiva_nome') or 'Sem perspectiva'
                o_id = link.get('objetivo_id')
                o_nome = link.get('objetivo_nome') or 'Sem objetivo'
                if p_id:
                    if p_id not in pdi_perspectiva_map:
                        pdi_perspectiva_map[p_id] = {'nome': p_nome, 'total': Decimal('0'), 'objetivos': {}}
                    pdi_perspectiva_map[p_id]['total'] += share
                    if o_id:
                        obj_map = pdi_perspectiva_map[p_id]['objetivos']
                        if o_id not in obj_map:
                            obj_map[o_id] = {'nome': o_nome, 'total': Decimal('0')}
                        obj_map[o_id]['total'] += share
                else:
                    sem_vinculo_pdi += share
        elif despesa.perspectiva_pdi_id:
            p = despesa.perspectiva_pdi
            if p.pk not in pdi_perspectiva_map:
                pdi_perspectiva_map[p.pk] = {'nome': p.nome, 'total': Decimal('0'), 'objetivos': {}}
            pdi_perspectiva_map[p.pk]['total'] += valor
            if despesa.objetivo_pdi_id:
                o = despesa.objetivo_pdi
                obj_map = pdi_perspectiva_map[p.pk]['objetivos']
                if o.pk not in obj_map:
                    obj_map[o.pk] = {'nome': o.nome, 'total': Decimal('0')}
                obj_map[o.pk]['total'] += valor
        else:
            sem_vinculo_pdi += valor

    pdi_perspectivas_sorted = sorted(pdi_perspectiva_map.items(), key=lambda x: -float(x[1]['total']))
    dados_pdi = json.dumps({
        'perspectivas': [
            {
                'id': p_id,
                'nome': p['nome'],
                'total': float(p['total']),
                'objetivos': sorted(
                    [{'id': o_id, 'nome': o['nome'], 'total': float(o['total'])} for o_id, o in p['objetivos'].items()],
                    key=lambda x: -x['total'],
                ),
            }
            for p_id, p in pdi_perspectivas_sorted
        ],
        'sem_vinculo': float(sem_vinculo_pdi),
    })

    # — Tabela cruzada: Origem × Ano —
    tc_anos = sorted({r.ano_fiscal for r in recursos_lista})
    tc_map = defaultdict(lambda: defaultdict(float))
    for recurso in recursos_lista:
        tc_map[recurso.origem_recurso or 'Sem origem'][recurso.ano_fiscal] += float(
            saldos_recursos.get(recurso.pk, Decimal('0'))
        )

    tc_linhas = []
    for orig, ano_vals in sorted(tc_map.items()):
        valores = [ano_vals.get(a, 0) for a in tc_anos]
        tc_linhas.append({'origem': orig, 'valores': valores, 'total': sum(valores)})

    tc_totais = [sum(l['valores'][i] for l in tc_linhas) for i in range(len(tc_anos))]
    tabela_cruzada = {
        'anos':       tc_anos,
        'linhas':     tc_linhas,
        'totais':     tc_totais,
        'grand_total': sum(tc_totais),
    }

    # — Tabela bens —
    tabela_recursos = [
        {
            'ano': recurso.ano_fiscal,
            'setor': _dashboard_setor_label(recurso.setor),
            'origem': recurso.origem_recurso,
            'natureza': recurso.natureza,
            'rubrica': recurso.rubrica,
            'orcado': recurso.valor_orcamentario,
            'saldo': saldos_recursos.get(recurso.pk, Decimal('0')),
        }
        for recurso in sorted(
            recursos_lista,
            key=lambda r: (r.ano_fiscal, _dashboard_setor_label(r.setor), r.origem_recurso, r.natureza, r.rubrica),
        )
    ]

    tabela_bens = list(
        despesas_qs.order_by('-valor_comprometido')
        .values('discriminacao', 'nota_empenho', 'quantidade', 'valor_comprometido')[:200]
    )

    # — Opções de filtro —
    origens_disp = list(
        RecursoOrcamentario.objects.values_list('origem_recurso', flat=True).distinct().order_by('origem_recurso')
    )
    naturezas_cadastradas = [
        n.nome
        for n in NaturezaRecurso.objects.filter(ativo=True).order_by('ordem', 'nome')
    ]
    naturezas_usadas = {
        str(n).strip()
        for n in RecursoOrcamentario.objects.exclude(natureza='')
        .values_list('natureza', flat=True)
        if str(n).strip()
    } | {
        str(n).strip()
        for n in Despesa.objects.exclude(natureza='')
        .values_list('natureza', flat=True)
        if str(n).strip()
    }
    naturezas_disp = sorted(
        set(naturezas_cadastradas) | naturezas_usadas,
        key=lambda n: (_dashboard_natureza_key(n), n),
    )
    rubricas_cadastradas = [
        str(r)
        for r in Rubrica.objects.filter(ativo=True).order_by('ordem', 'codigo')
    ]
    rubricas_usadas = set(
        RecursoOrcamentario.objects.exclude(rubrica='')
        .values_list('rubrica', flat=True)
    ) | set(
        Despesa.objects.exclude(rubrica='')
        .values_list('rubrica', flat=True)
    )
    rubricas_disp = [(r, r) for r in sorted(set(rubricas_cadastradas) | rubricas_usadas)]
    setores_disp = (
        Setor.objects
        .filter(ativo=True, recursos_orcamentarios__isnull=False)
        .distinct()
        .order_by('sigla')
    )

    ctx = {
        'kpi':             kpi,
        'dados_area':      dados_area,
        'dados_setor':     dados_setor,
        'dados_rubrica':   dados_rubrica,
        'dados_pdi':       dados_pdi,
        'tabela_cruzada':  tabela_cruzada,
        'tabela_recursos': tabela_recursos,
        'tabela_bens':     tabela_bens,
        'total_orcado_json': json.dumps(float(total_orcado)),
        'total_comprometido_json': json.dumps(float(total_comprometido)),
        'filtros':         filtros,
        'anos_selecionados':  [str(a) for a in anos_selecionados],
        'anos_disponiveis':   _anos_disponiveis(),
        'naturezas':          naturezas_disp,
        'rubricas':           rubricas_disp,
        'origens':            origens_disp,
        'setores':            setores_disp,
    }
    return render(request, 'orcamento/dashboard.html', ctx)


# ── Painel interno ────────────────────────────────────────────────────────────

@gestor_financeiro_required
def orcamento_home(request):
    ano_atual = timezone.now().year
    recursos = RecursoOrcamentario.objects.filter(ano_fiscal=ano_atual).select_related('setor')
    total_orcado = recursos.aggregate(t=Sum('valor_orcamentario'))['t'] or 0

    despesas_ano = Despesa.objects.filter(
        recurso__ano_fiscal=ano_atual
    ).exclude(situacao__in=SituacaoDespesa.chaves_sem_impacto())
    total_comprometido = despesas_ano.aggregate(t=Sum('valor_comprometido'))['t'] or 0

    saldo_disponivel = float(total_orcado) - float(total_comprometido)
    pct_execucao = (float(total_comprometido) / float(total_orcado) * 100) if total_orcado else 0

    ultimas_despesas = Despesa.objects.select_related('setor').order_by('-criada_em')[:10]

    ctx = {
        'total_orcado':          total_orcado,
        'total_comprometido':    total_comprometido,
        'saldo_disponivel':      saldo_disponivel,
        'pct_execucao':          pct_execucao,
        'ultimas_despesas':      ultimas_despesas,
    }
    return render(request, 'orcamento/home.html', ctx)


# ── PDI ───────────────────────────────────────────────────────────────────────

def _pdi_htmx(request):
    return request.headers.get('HX-Request') == 'true'


def _pdi_modal_render(request, form, title, fallback_tpl, fallback_ctx=None):
    ctx = {'form': form, 'modal_title': title, 'form_action': request.path}
    if _pdi_htmx(request):
        return render(request, 'orcamento/pdi/_modal_form.html', ctx)
    return render(request, fallback_tpl, fallback_ctx or ctx)


def _pdi_done(request, msg):
    messages.success(request, msg)
    if _pdi_htmx(request):
        r = HttpResponse()
        r['HX-Redirect'] = reverse('pdi_perspectiva_list')
        return r
    return redirect('pdi_perspectiva_list')


@gestor_financeiro_required
def pdi_perspectiva_list(request):
    perspectivas = PdiPerspectiva.objects.prefetch_related(
        'objetivos__indicadores'
    ).order_by('ordem', 'nome')
    return render(request, 'orcamento/pdi/perspectiva_list.html', {'perspectivas': perspectivas})


@gestor_financeiro_required
def pdi_perspectiva_create(request):
    form = PdiPerspectivaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return _pdi_done(request, 'Perspectiva criada.')
    return _pdi_modal_render(request, form, 'Nova Perspectiva',
                             'orcamento/pdi/perspectiva_form.html')


@gestor_financeiro_required
def pdi_perspectiva_edit(request, pk):
    obj = get_object_or_404(PdiPerspectiva, pk=pk)
    form = PdiPerspectivaForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return _pdi_done(request, 'Perspectiva atualizada.')
    return _pdi_modal_render(request, form, 'Editar Perspectiva',
                             'orcamento/pdi/perspectiva_form.html', {'form': form, 'object': obj})


@gestor_financeiro_required
def pdi_perspectiva_delete(request, pk):
    obj = get_object_or_404(PdiPerspectiva, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Perspectiva removida.')
    return redirect('pdi_perspectiva_list')


@gestor_financeiro_required
def pdi_objetivo_create(request):
    perspectiva_pk = request.POST.get('perspectiva') or request.GET.get('perspectiva')
    initial = {'perspectiva': perspectiva_pk} if perspectiva_pk else {}
    form = PdiObjetivoForm(request.POST or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return _pdi_done(request, 'Objetivo estratégico criado.')
    return _pdi_modal_render(request, form, 'Novo Objetivo Estratégico',
                             'orcamento/pdi/objetivo_form.html')


@gestor_financeiro_required
def pdi_objetivo_edit(request, pk):
    obj = get_object_or_404(PdiObjetivoEstrategico, pk=pk)
    form = PdiObjetivoForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return _pdi_done(request, 'Objetivo atualizado.')
    return _pdi_modal_render(request, form, 'Editar Objetivo Estratégico',
                             'orcamento/pdi/objetivo_form.html', {'form': form, 'object': obj})


@gestor_financeiro_required
def pdi_objetivo_delete(request, pk):
    obj = get_object_or_404(PdiObjetivoEstrategico, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Objetivo removido.')
    return redirect('pdi_perspectiva_list')


@gestor_financeiro_required
def pdi_indicador_create(request):
    objetivo_pk = request.POST.get('objetivo') or request.GET.get('objetivo')
    initial = {'objetivo': objetivo_pk} if objetivo_pk else {}
    form = PdiIndicadorForm(request.POST or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return _pdi_done(request, 'Indicador criado.')
    return _pdi_modal_render(request, form, 'Novo Indicador',
                             'orcamento/pdi/indicador_form.html')


@gestor_financeiro_required
def pdi_indicador_edit(request, pk):
    obj = get_object_or_404(PdiIndicador, pk=pk)
    form = PdiIndicadorForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return _pdi_done(request, 'Indicador atualizado.')
    return _pdi_modal_render(request, form, 'Editar Indicador',
                             'orcamento/pdi/indicador_form.html', {'form': form, 'object': obj})


@gestor_financeiro_required
def pdi_indicador_delete(request, pk):
    obj = get_object_or_404(PdiIndicador, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Indicador removido.')
    return redirect('pdi_perspectiva_list')


@gestor_financeiro_required
def pdi_importar(request):
    if request.method != 'POST':
        return redirect('pdi_perspectiva_list')

    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        messages.error(request, 'Nenhum arquivo enviado.')
        return redirect('pdi_perspectiva_list')

    try:
        import openpyxl

        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb.active

        headers = [
            (str(c.value).strip().lower() if c.value else '') for c in ws[1]
        ]

        def _col(keywords):
            for i, h in enumerate(headers):
                if any(k in h for k in keywords):
                    return i
            return None

        ci_p = _col(['perspectiva'])
        ci_o = _col(['objetivo'])
        ci_i = _col(['indicador'])

        if ci_p is None:
            messages.error(request, 'Coluna "Perspectiva" não encontrada na planilha.')
            return redirect('pdi_perspectiva_list')

        criados = [0, 0, 0]

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                p_val = str(row[ci_p]).strip() if ci_p is not None and row[ci_p] else ''
                o_val = str(row[ci_o]).strip() if ci_o is not None and row[ci_o] else ''
                i_val = str(row[ci_i]).strip() if ci_i is not None and row[ci_i] else ''

                if not p_val or p_val.lower() == 'none':
                    continue

                p_obj, p_new = PdiPerspectiva.objects.get_or_create(nome=p_val)
                if p_new:
                    criados[0] += 1

                if not o_val or o_val.lower() == 'none':
                    continue

                oe_obj, oe_new = PdiObjetivoEstrategico.objects.get_or_create(
                    perspectiva=p_obj, nome=o_val
                )
                if oe_new:
                    criados[1] += 1

                if not i_val or i_val.lower() == 'none':
                    continue

                _, ind_new = PdiIndicador.objects.get_or_create(
                    objetivo=oe_obj, nome=i_val
                )
                if ind_new:
                    criados[2] += 1

        messages.success(
            request,
            f'Importação concluída: {criados[0]} perspectiva(s), '
            f'{criados[1]} objetivo(s) e {criados[2]} indicador(es) criados.'
        )
    except Exception as exc:
        messages.error(request, f'Erro na importação: {exc}')

    return redirect('pdi_perspectiva_list')


@gestor_financeiro_required
def pdi_modelo_xlsx(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PDI'

    FILL_H = PatternFill('solid', fgColor='C6543C')
    FONT_H = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    FONT_B = Font(name='Calibri', size=10)
    AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    AL_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Perspectiva', 'Objetivo Estratégico', 'Indicador']
    widths = [32, 78, 78]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = FONT_H
        cell.fill = FILL_H
        cell.alignment = AL_C
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    sample = [
        ('Eficiência Orçamentária',
         'OE.01 - Assegurar uma gestão orçamentária e financeira eficiente, '
         'buscando o alinhamento entre orçamento, estratégia e gestão',
         'OE.01.I1 - Taxa de planejamento orçamentário (PO) alinhado ao PDI (TPO)'),
        ('Eficiência Orçamentária',
         'OE.01 - Assegurar uma gestão orçamentária e financeira eficiente, '
         'buscando o alinhamento entre orçamento, estratégia e gestão',
         'OE.01.I2 - Taxa de execução orçamentária (TEO)'),
        ('Eficiência Orçamentária',
         'OE.02 - Ampliar a captação de recursos externos',
         'OE.02.I1 - Número de parcerias com recursos captados destinados a projetos de desenvolvimento institucional'),
        ('Aprendizado e Crescimento',
         'OE.03 - Promover capacitação e qualificação dos servidores com destaque para as áreas estratégicas',
         'OE.03.I1 - Taxa de desenvolvimento de competências (TDC)'),
        ('Aprendizado e Crescimento',
         'OE.03 - Promover capacitação e qualificação dos servidores com destaque para as áreas estratégicas',
         'OE.03.I2 - Número de certificações em cursos de capacitação'),
        ('Processos Internos',
         'OE.11 - Aprimorar práticas administrativas baseadas nos princípios da boa governança e gestão pública',
         'OE.11.I1 - Taxa de participação da comunidade acadêmica nos instrumentos de autoavaliação institucional (TPCAS)'),
        ('UFPB e Sociedade',
         'OE.17 - Fortalecer o desempenho institucional e ampliar seu impacto na sociedade',
         'OE.17.I1 - Conceito preliminar de curso com ENADE'),
    ]

    FILLS_ROW = [PatternFill('solid', fgColor='FDF0ED'), PatternFill('solid', fgColor='FFFFFF')]
    for r_idx, row in enumerate(sample, 2):
        fill = FILLS_ROW[r_idx % 2]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font = FONT_B
            cell.alignment = AL_L
            cell.border = BORDER
            cell.fill = fill
        ws.row_dimensions[r_idx].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="PDI_modelo_importacao.xlsx"'
    return resp


# ── Recursos Orçamentários ────────────────────────────────────────────────────

@gestor_financeiro_required
def recurso_list(request):
    SORT_MAP = {
        'ano':      'ano_fiscal',
        'setor':    'setor__sigla',
        'origem':   'origem_recurso',
        'natureza': 'natureza',
        'rubrica':  'rubrica',
        'orcado':   'valor_orcamentario',
        'saldo':    None,
    }
    COLUNAS = [
        ('ano',      'Ano',      'px-3'),
        ('setor',    'Setor',    ''),
        ('origem',   'Origem',   ''),
        ('natureza', 'Natureza', ''),
        ('rubrica',  'Rubrica',  ''),
        ('orcado',   'Orçado',   'text-end'),
        ('saldo',    'Saldo',    'text-end'),
    ]

    sort = request.GET.get('sort', 'ano')
    dir_ = request.GET.get('dir', 'desc')
    if sort not in SORT_MAP:
        sort = 'ano'

    # Filtros multi-select
    sel_anos      = request.GET.getlist('ano')
    sel_setores   = request.GET.getlist('setor')
    sel_origens   = request.GET.getlist('origem')
    sel_naturezas = request.GET.getlist('natureza')
    sel_rubricas  = request.GET.getlist('rubrica')

    # Dataset completo para cascade (antes dos filtros)
    all_qs = RecursoOrcamentario.objects.select_related('setor').all()
    all_records_json = json.dumps([
        {
            'ano':      str(r.ano_fiscal),
            'setor':    str(r.setor_id) if r.setor_id else '',
            'origem':   r.origem_recurso,
            'natureza': r.natureza,
            'rubrica':  r.rubrica,
        }
        for r in all_qs
    ])

    # Queryset filtrado
    qs = all_qs
    if sel_anos:      qs = qs.filter(ano_fiscal__in=sel_anos)
    if sel_setores:
        setor_ids = [s for s in sel_setores if str(s).isdigit()]
        setor_siglas = [s for s in sel_setores if not str(s).isdigit()]
        setor_filter = Q()
        if setor_ids:
            setor_filter |= Q(setor_id__in=setor_ids)
        if setor_siglas:
            setor_filter |= Q(setor__sigla__in=setor_siglas)
        qs = qs.filter(setor_filter)
    if sel_origens:   qs = qs.filter(origem_recurso__in=sel_origens)
    if sel_naturezas: qs = qs.filter(natureza__in=sel_naturezas)
    if sel_rubricas:  qs = qs.filter(rubrica__in=sel_rubricas)

    # Ordenação
    orm_field = SORT_MAP[sort]
    if orm_field:
        qs = qs.order_by(f'-{orm_field}' if dir_ == 'desc' else orm_field)
    else:
        qs = qs.order_by('-ano_fiscal', 'setor__sigla')

    total_orcado_bruto = qs.aggregate(t=Sum('valor_orcamentario'))['t'] or 0
    recursos = list(qs)
    total_saldo = sum(r.saldo_atual for r in recursos)
    # Ajuste por transferências (entradas - saídas) dentro do conjunto filtrado
    _rec_ids = [r.pk for r in recursos]
    _entradas = Transferencia.objects.filter(status='realizada', destino_id__in=_rec_ids).aggregate(t=Sum('valor'))['t'] or 0
    _saidas = Transferencia.objects.filter(status='realizada', origem_id__in=_rec_ids).aggregate(t=Sum('valor'))['t'] or 0
    total_orcado = total_orcado_bruto + _entradas - _saidas
    if sort == 'saldo':
        recursos.sort(key=lambda r: r.saldo_atual, reverse=(dir_ == 'desc'))

    # ── Linhas virtuais de transferências (débito/crédito) ───────────────────
    from types import SimpleNamespace
    def _rec_passa(rec):
        if sel_anos and str(rec.ano_fiscal) not in sel_anos:
            return False
        if sel_setores:
            ids = [s for s in sel_setores if str(s).isdigit()]
            sig = [s for s in sel_setores if not str(s).isdigit()]
            ok = (ids and rec.setor_id and rec.setor_id in [int(i) for i in ids]) \
                 or (sig and rec.setor and rec.setor.sigla in sig)
            if not ok:
                return False
        if sel_origens and rec.origem_recurso not in sel_origens:
            return False
        if sel_naturezas and rec.natureza not in sel_naturezas:
            return False
        if sel_rubricas and str(rec.rubrica) not in sel_rubricas:
            return False
        return True

    def _linha_transf(rec, valor, tipo, t):
        return SimpleNamespace(
            pk=rec.pk,
            ano_fiscal=rec.ano_fiscal,
            setor=rec.setor,
            origem_recurso=rec.origem_recurso,
            natureza=rec.natureza,
            rubrica=rec.rubrica,
            valor_orcamentario=valor,
            saldo_atual=None,
            is_transferencia=True,
            transf_tipo=tipo,
            transf_id=t.pk,
            transf_data=t.data,
        )

    transf_qs = (Transferencia.objects
                 .select_related('origem__setor', 'destino__setor')
                 .filter(status='realizada'))
    linhas_transf = []
    for t in transf_qs:
        if _rec_passa(t.origem):
            linhas_transf.append(_linha_transf(t.origem, -t.valor, 'debito', t))
        if _rec_passa(t.destino):
            linhas_transf.append(_linha_transf(t.destino, t.valor, 'credito', t))
    linhas_transf.sort(key=lambda r: (r.ano_fiscal, r.transf_data or 0), reverse=True)


    # Mapeamento estático Natureza ↔ Rubrica (para cascade mesmo sem registros)
    rubrica_qs = Rubrica.objects.filter(ativo=True).select_related('natureza').order_by('ordem', 'codigo')
    rubrica_map = {}        # natureza_nome -> [rubrica_str, ...]
    natureza_de_rubrica = {}  # rubrica_str -> natureza_nome
    for rb in rubrica_qs:
        nat = rb.natureza.nome if rb.natureza else ''
        rb_str = str(rb)
        rubrica_map.setdefault(nat, []).append(rb_str)
        natureza_de_rubrica[rb_str] = nat

    ctx = {
        'recursos':               recursos,
        'linhas_transf':          linhas_transf,
        'all_records_json':       all_records_json,
        'rubrica_map_json':       json.dumps(rubrica_map),
        'natureza_de_rubrica_json': json.dumps(natureza_de_rubrica),
        'anos_opts':              _anos_disponiveis(),
        'setores_opts':           Setor.objects.filter(ativo=True).order_by('sigla', 'codigo'),
        'origens_opts':           OrigemRecurso.objects.filter(ativo=True).order_by('ordem', 'nome'),
        'naturezas_opts':         NaturezaRecurso.objects.filter(ativo=True).order_by('ordem', 'nome'),
        'rubricas_opts':          list(rubrica_qs),
        'sel_anos':          sel_anos,
        'sel_setores':       sel_setores,
        'sel_origens':       sel_origens,
        'sel_naturezas':     sel_naturezas,
        'sel_rubricas':      sel_rubricas,
        'total_orcado':      total_orcado,
        'total_saldo':       total_saldo,
        'sort':              sort,
        'dir':               dir_,
        'colunas':           COLUNAS,
        'filtros_ativos':    any([sel_anos, sel_setores, sel_origens, sel_naturezas, sel_rubricas]),
    }
    return render(request, 'orcamento/recurso_list.html', ctx)


@gestor_financeiro_required
def recurso_template_xlsx(request):
    """Gera e devolve a planilha modelo para importação de recursos orçamentários."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    setores = list(Setor.objects.filter(ativo=True).order_by('sigla', 'codigo'))
    origens = list(OrigemRecurso.objects.filter(ativo=True).order_by('ordem', 'nome'))
    naturezas = list(NaturezaRecurso.objects.filter(ativo=True).order_by('ordem', 'nome'))
    rubricas_qs = list(Rubrica.objects.filter(ativo=True).select_related('natureza').order_by('ordem', 'codigo'))

    setor_labels = [_setor_catalog_label(s) for s in setores]
    origem_labels = [o.nome for o in origens]
    natureza_labels = [n.nome for n in naturezas]
    rubrica_labels = [str(r) for r in rubricas_qs]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Recursos Orçamentários'

    headers = [
        'ANO FISCAL', 'SETOR', 'ORIGEM DO RECURSO',
        'NATUREZA', 'RUBRICA', 'VALOR ORÇAMENTÁRIO', 'OBSERVAÇÕES',
    ]
    col_widths = [12, 42, 35, 18, 42, 22, 40]

    header_fill = PatternFill('solid', fgColor='C6543C')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # Linha de exemplo
    exemplo_natureza = natureza_labels[0] if natureza_labels else ''
    exemplo_rubrica = next(
        (str(r) for r in rubricas_qs if r.natureza and r.natureza.nome == exemplo_natureza),
        rubrica_labels[0] if rubrica_labels else '',
    )
    example = [
        timezone.now().year,
        setor_labels[0] if setor_labels else '',
        origem_labels[0] if origem_labels else '',
        exemplo_natureza,
        exemplo_rubrica,
        '50000,00',
        '',
    ]
    ex_font = Font(name='Calibri', size=10, italic=True, color='555555')
    for col_idx, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = ex_font
        cell.border = border

    # Aba de referência: valores válidos
    ws_ref = wb.create_sheet('Referência')
    ref_data = [
        ('SETOR', setor_labels, 42),
        ('ORIGEM DO RECURSO', origem_labels, 35),
        ('NATUREZA', natureza_labels, 18),
        ('RUBRICA', rubrica_labels, 42),
    ]
    for col_idx, (title, values, width) in enumerate(ref_data, 1):
        cell = ws_ref.cell(row=1, column=col_idx, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws_ref.column_dimensions[cell.column_letter].width = width
        for row_idx, value in enumerate(values, 2):
            ref_cell = ws_ref.cell(row=row_idx, column=col_idx, value=value)
            ref_cell.border = border

    def add_list_validation(col_letter, ref_col, values_count, label):
        if values_count <= 0:
            return
        formula = f"'Referência'!${ref_col}$2:${ref_col}${values_count + 1}"
        dv = DataValidation(type='list', formula1=formula, allow_blank=False)
        dv.errorTitle = f'{label} inválido'
        dv.error = f'Escolha um valor cadastrado na aba Referência para {label}.'
        dv.promptTitle = label
        dv.prompt = 'Selecione um valor da lista.'
        dv.showErrorMessage = True
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}501')

    add_list_validation('B', 'A', len(setor_labels), 'SETOR')
    add_list_validation('C', 'B', len(origem_labels), 'ORIGEM DO RECURSO')
    add_list_validation('D', 'C', len(natureza_labels), 'NATUREZA')
    add_list_validation('E', 'D', len(rubrica_labels), 'RUBRICA')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:G1'
    ws_ref.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_recursos_orcamentarios.xlsx"'
    wb.save(response)
    return response


# ── Helpers de importação ─────────────────────────────────────────────────────

def _orc_norm(s):
    if s is None:
        return ''
    s = str(s).strip().upper()
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def _orc_parse_decimal(val):
    if val is None:
        return None
    s = str(val).replace('R$', '').replace('\xa0', '').strip()
    if not s:
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _orc_rows_from_xlsx(arquivo):
    import openpyxl
    wb = openpyxl.load_workbook(filename=io.BytesIO(arquivo.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_orc_norm(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _orc_rows_from_csv(arquivo):
    decoded = arquivo.read().decode('utf-8-sig')
    reader = csv.DictReader(decoded.splitlines(), delimiter=';')
    return [{_orc_norm(k): v for k, v in row.items()} for row in reader]


def _resolver_setor(raw):
    """Resolve setor por rótulo do catálogo, código SIPAC ou sigla."""
    if not raw:
        return None
    texto = str(raw).strip()
    texto_norm = _orc_norm(texto)

    codigo = ''
    if texto.endswith(')') and '(' in texto:
        codigo = texto.rsplit('(', 1)[1].rstrip(')').strip()

    candidatos = Setor.objects.filter(ativo=True)
    if codigo:
        setor = candidatos.filter(codigo__iexact=codigo).first()
        if setor:
            return setor

    setor = candidatos.filter(codigo__iexact=texto).first()
    if setor:
        return setor

    setor = candidatos.filter(sigla__iexact=texto).first()
    if setor:
        return setor

    for sep in (' — ', ' - '):
        if sep in texto:
            sigla = texto.split(sep, 1)[0].strip()
            setor = candidatos.filter(sigla__iexact=sigla).first()
            if setor:
                return setor

    for setor in candidatos:
        if texto_norm in {
            _orc_norm(_setor_catalog_label(setor)),
            _orc_norm(setor.nome),
            _orc_norm(setor.sigla),
            _orc_norm(setor.codigo),
        }:
            return setor
    return None


def _resolver_origem(raw):
    """Resolve origem do recurso a partir do cadastro ativo."""
    if not raw:
        return ''
    raw_norm = _orc_norm(str(raw))
    for origem in OrigemRecurso.objects.filter(ativo=True):
        nome_norm = _orc_norm(origem.nome)
        if raw_norm == nome_norm or raw_norm in nome_norm or nome_norm in raw_norm:
            return origem.nome
    return ''


def _resolver_rubrica(raw):
    """Resolve o valor da rubrica a partir do texto da planilha.
    Retorna o label completo 'CODIGO - Nome' conforme cadastrado, ou string vazia."""
    if not raw:
        return ''
    raw_norm = _orc_norm(str(raw))
    # Código numérico extraído do início (ex: "339030")
    code_raw = str(raw).strip().split(' ')[0].strip()
    raw_tem_nome = any(ch.isalpha() for ch in raw_norm)
    matches_codigo = []

    for r in Rubrica.objects.filter(ativo=True):
        label = str(r)           # "339030 - Material de Consumo"
        label_norm = _orc_norm(label)
        if raw_norm == label_norm:
            return label
        if raw_tem_nome and (raw_norm in label_norm or label_norm in raw_norm):
            return label
        if code_raw == r.codigo:
            matches_codigo.append(label)
    if len(matches_codigo) == 1:
        return matches_codigo[0]
    return ''


def _resolver_natureza(raw):
    """Resolve o valor da natureza a partir do texto da planilha.
    Retorna o nome conforme cadastrado em NaturezaRecurso."""
    if not raw:
        return ''
    raw_norm = _orc_norm(str(raw))
    for n in NaturezaRecurso.objects.filter(ativo=True):
        if raw_norm == _orc_norm(n.nome):
            return n.nome
        if raw_norm in _orc_norm(n.nome) or _orc_norm(n.nome) in raw_norm:
            return n.nome
    return ''


@gestor_financeiro_required
def recurso_importar(request):
    """Importa recursos orçamentários via planilha Excel ou CSV."""
    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            messages.error(request, 'Nenhum arquivo enviado.')
            return redirect('recurso_importar')

        nome = arquivo.name.lower()
        try:
            if nome.endswith('.xlsx'):
                linhas = _orc_rows_from_xlsx(arquivo)
            else:
                linhas = _orc_rows_from_csv(arquivo)
        except Exception as e:
            messages.error(request, f'Erro ao ler arquivo: {e}')
            return redirect('recurso_importar')

        count = 0
        erros = []
        rubrica_natureza_por_label = {
            str(r): (r.natureza.nome if r.natureza else '')
            for r in Rubrica.objects.filter(ativo=True).select_related('natureza')
        }

        from django.db import transaction
        with transaction.atomic():
            for i, row in enumerate(linhas, 2):  # linha 2 = após cabeçalho
                try:
                    ano_raw = row.get(_orc_norm('ANO FISCAL')) or row.get(_orc_norm('ANO'))
                    try:
                        ano = int(str(ano_raw).strip())
                    except (ValueError, TypeError):
                        erros.append(f'Linha {i}: ANO FISCAL inválido ("{ano_raw}").')
                        continue

                    origem_raw = (
                        row.get(_orc_norm('ORIGEM DO RECURSO')) or
                        row.get(_orc_norm('ORIGEM')) or ''
                    )
                    origem = _resolver_origem(origem_raw)
                    if not origem:
                        erros.append(f'Linha {i}: ORIGEM DO RECURSO inválida ou não cadastrada ("{origem_raw}").')
                        continue

                    natureza = _resolver_natureza(
                        row.get(_orc_norm('NATUREZA')) or
                        row.get(_orc_norm('NATUREZA DO RECURSO'))
                    )
                    rubrica = _resolver_rubrica(
                        row.get(_orc_norm('RUBRICA'))
                    )
                    if not natureza:
                        erros.append(f'Linha {i}: NATUREZA inválida ou não cadastrada.')
                        continue
                    if not rubrica:
                        erros.append(f'Linha {i}: RUBRICA inválida ou não cadastrada.')
                        continue
                    rubrica_natureza = rubrica_natureza_por_label.get(rubrica, '')
                    if rubrica_natureza and rubrica_natureza != natureza:
                        erros.append(
                            f'Linha {i}: RUBRICA "{rubrica}" pertence à natureza '
                            f'"{rubrica_natureza}", não "{natureza}".'
                        )
                        continue
                    valor = _orc_parse_decimal(
                        row.get(_orc_norm('VALOR ORCAMENTARIO')) or
                        row.get(_orc_norm('VALOR'))
                    )
                    if valor is None or valor <= 0:
                        erros.append(f'Linha {i}: VALOR ORÇAMENTÁRIO inválido.')
                        continue

                    observacoes = str(
                        row.get(_orc_norm('OBSERVACOES')) or
                        row.get(_orc_norm('OBS')) or ''
                    ).strip()

                    setor_raw = (
                        row.get(_orc_norm('SETOR (SIGLA)')) or
                        row.get(_orc_norm('SETOR')) or ''
                    )
                    setor = _resolver_setor(setor_raw)
                    if not setor:
                        erros.append(f'Linha {i}: SETOR inválido ou não cadastrado ("{setor_raw}").')
                        continue

                    obj, created = RecursoOrcamentario.objects.update_or_create(
                        ano_fiscal=ano,
                        setor=setor,
                        origem_recurso=origem,
                        natureza=natureza,
                        rubrica=rubrica,
                        defaults={
                            'valor_orcamentario': valor,
                            'observacoes': observacoes,
                            'criado_por': request.user,
                        },
                    )
                    count += 1
                except Exception as e:
                    erros.append(f'Linha {i}: {e}')

        if erros:
            messages.warning(
                request,
                f'{count} recurso(s) importado(s). {len(erros)} erro(s): ' +
                ' | '.join(erros[:5])
            )
        else:
            messages.success(request, f'{count} recurso(s) importado(s) com sucesso!')
        return redirect('recurso_list')

    return render(request, 'orcamento/recurso_importar.html')


def _rubrica_map_ctx():
    """Contexto com mapeamento Natureza↔Rubrica para cascade no formulário."""
    rubrica_map = {}
    natureza_de_rubrica = {}
    for rb in Rubrica.objects.filter(ativo=True).select_related('natureza').order_by('ordem', 'codigo'):
        nat = rb.natureza.nome if rb.natureza else ''
        rb_str = str(rb)
        rubrica_map.setdefault(nat, []).append(rb_str)
        natureza_de_rubrica[rb_str] = nat
    return {
        'rubrica_map_json':          json.dumps(rubrica_map),
        'natureza_de_rubrica_json':  json.dumps(natureza_de_rubrica),
    }


@gestor_financeiro_required
def recurso_create(request):
    form = RecursoOrcamentarioForm(request.POST or None)
    if form.is_valid():
        r = form.save(commit=False)
        r.criado_por = request.user
        r.save()
        messages.success(request, 'Recurso orçamentário criado.')
        return redirect('recurso_list')
    ctx = {'form': form, **_rubrica_map_ctx()}
    return render(request, 'orcamento/recurso_form.html', ctx)


@gestor_financeiro_required
def recurso_edit(request, pk):
    obj = get_object_or_404(RecursoOrcamentario, pk=pk)
    form = RecursoOrcamentarioForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Recurso atualizado.')
        return redirect('recurso_list')
    ctx = {'form': form, 'object': obj, **_rubrica_map_ctx()}
    return render(request, 'orcamento/recurso_form.html', ctx)


@gestor_financeiro_required
def recurso_delete(request, pk):
    obj = get_object_or_404(RecursoOrcamentario, pk=pk)
    if request.method == 'POST':
        if obj.despesas.exists():
            messages.error(request, 'Não é possível excluir: há despesas vinculadas a este recurso.')
            return redirect('recurso_list')
        obj.delete()
        messages.success(request, 'Recurso excluído.')
        return redirect('recurso_list')
    return render(request, 'orcamento/recurso_confirm_delete.html', {'object': obj})


@gestor_financeiro_required
def recurso_excluir_lote(request):
    if request.method != 'POST':
        return redirect('recurso_list')
    pks = request.POST.getlist('pks')
    if not pks:
        messages.warning(request, 'Nenhum recurso selecionado.')
        return redirect('recurso_list')
    qs = RecursoOrcamentario.objects.filter(pk__in=pks)
    com_despesas = qs.filter(despesas__isnull=False).distinct()
    if com_despesas.exists():
        nomes = ', '.join(str(r) for r in com_despesas[:5])
        messages.error(request, f'Não foi possível excluir {com_despesas.count()} recurso(s) com despesas vinculadas: {nomes}.')
        qs = qs.exclude(pk__in=com_despesas.values('pk'))
    excluidos = qs.count()
    qs.delete()
    if excluidos:
        messages.success(request, f'{excluidos} recurso(s) excluído(s) com sucesso.')
    return redirect('recurso_list')


# ── Transferências ────────────────────────────────────────────────────────────

@gestor_financeiro_required
def transferencia_list(request):
    qs = Transferencia.objects.select_related(
        'origem__setor', 'destino__setor', 'criada_por'
    )
    status = request.GET.get('status', '')
    if status:
        qs = qs.filter(status=status)

    # Ordenação por coluna
    ORDEM_MAP = {
        'data':              'data',
        '-data':             '-data',
        'valor':             'valor',
        '-valor':            '-valor',
        'origem_setor':      'origem__setor__sigla',
        '-origem_setor':     '-origem__setor__sigla',
        'destino_setor':     'destino__setor__sigla',
        '-destino_setor':    '-destino__setor__sigla',
        'status':            'status',
        '-status':           '-status',
        'criada_por':        'criada_por__first_name',
        '-criada_por':       '-criada_por__first_name',
        'criada_em':         'criada_em',
        '-criada_em':        '-criada_em',
    }
    ordem_param = request.GET.get('ordem', '-criada_em')
    ordem_db = ORDEM_MAP.get(ordem_param, '-criada_em')
    qs = qs.order_by(ordem_db)

    return render(request, 'orcamento/transferencia_list.html', {
        'transferencias': qs,
        'ordem': ordem_param,
    })


@gestor_financeiro_required
def transferencia_detail(request, pk):
    t = get_object_or_404(Transferencia, pk=pk)
    return render(request, 'orcamento/transferencia_detail.html', {'transferencia': t})


@gestor_financeiro_required
def transferencia_edit(request, pk):
    t = get_object_or_404(Transferencia, pk=pk)
    if t.status == 'cancelada':
        messages.error(request, 'Transferências canceladas não podem ser editadas.')
        return redirect('transferencia_list')

    ano_atual = timezone.now().year

    if request.method == 'POST':
        post_data = _resolver_destino(request.POST, request.user, ano_atual)
        form = TransferenciaForm(post_data, request.FILES or None, instance=t)
    else:
        form = TransferenciaForm(instance=t)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Transferência atualizada.')
        return redirect('transferencia_list')

    ctx = _contexto_transferencia(timezone.now().year)
    ctx['form']   = form
    ctx['object'] = t
    return render(request, 'orcamento/transferencia_form.html', ctx)


@gestor_financeiro_required
def transferencia_cancelar(request, pk):
    t = get_object_or_404(Transferencia, pk=pk)
    if request.method == 'POST':
        if t.status == 'cancelada':
            messages.warning(request, 'Transferência já está cancelada.')
        else:
            t.status = 'cancelada'
            t.save()
            messages.success(request, 'Transferência cancelada.')
        return redirect('transferencia_list')
    return render(request, 'orcamento/transferencia_cancelar_confirm.html', {'transferencia': t})


def _contexto_transferencia(ano_atual):
    """Monta o contexto comum às views de criação/edição de transferência."""
    from datetime import date

    recursos = (
        RecursoOrcamentario.objects
        .select_related('setor')
        .filter(ano_fiscal=ano_atual)
        .order_by('setor__sigla', 'origem_recurso', 'natureza', 'rubrica')
    )
    recursos_json = json.dumps([
        {
            'id': r.pk,
            'ano_fiscal': str(r.ano_fiscal),
            'setor': str(r.setor_id) if r.setor_id else '',
            'setor_label': _setor_catalog_label(r.setor),
            'origem_recurso': r.origem_recurso,
            'natureza': r.natureza,
            'rubrica': r.rubrica,
            'saldo_atual': float(r.saldo_atual),
        }
        for r in recursos
    ])

    # Todos os setores activos → DESTINO pode receber em qualquer setor
    setores_json = json.dumps([
        {
            'id': str(s.pk),
            'sigla': s.sigla,
            'nome': s.nome,
            'codigo': s.codigo,
            'label': _setor_catalog_label(s),
        }
        for s in Setor.objects.filter(ativo=True).order_by('sigla', 'codigo')
    ])

    # Todas as origens de recurso do catálogo
    origens_json = json.dumps([
        o.nome
        for o in OrigemRecurso.objects.filter(ativo=True).order_by('ordem', 'nome')
    ])

    naturezas_json = json.dumps([
        n.nome
        for n in NaturezaRecurso.objects.filter(ativo=True).order_by('ordem', 'nome')
    ])

    rubrica_map_transf = {}
    todas_rubricas = []
    for rb in Rubrica.objects.filter(ativo=True).select_related('natureza').order_by('ordem', 'codigo'):
        rb_str = str(rb)
        todas_rubricas.append(rb_str)
        if rb.natureza:
            nat_key = rb.natureza.nome
            rubrica_map_transf.setdefault(nat_key, []).append(rb_str)
            rubrica_map_transf.setdefault(nat_key.lower(), []).append(rb_str)

    campos_painel = [
        ('setor',          'Setor'),
        ('origem_recurso', 'Origem do Recurso'),
        ('natureza',       'Natureza'),
        ('rubrica',        'Rubrica'),
    ]

    return {
        'recursos_json':          recursos_json,
        'setores_json':           setores_json,
        'origens_json':           origens_json,
        'naturezas_json':         naturezas_json,
        'campos_painel':          campos_painel,
        'today_str':              date.today().isoformat(),
        'ano_atual':              ano_atual,
        'rubrica_map_transf_json': json.dumps(rubrica_map_transf),
        'todas_rubricas_json':    json.dumps(todas_rubricas),
    }


def _resolver_destino(post_data, user, ano_atual):
    """
    Recebe o QueryDict do POST, tenta obter ou criar o RecursoOrcamentario
    de destino a partir dos campos individuais enviados pelo formulário.
    Retorna o post_data modificado com 'destino' preenchido, ou o original
    se não for possível resolver.
    """
    setor_ref = post_data.get('destino_setor', '').strip()
    natureza  = post_data.get('destino_natureza', '').strip()
    rubrica   = post_data.get('destino_rubrica', '').strip()

    origem_pk = post_data.get('origem', '').strip()
    origem = ''
    if origem_pk:
        try:
            origem = RecursoOrcamentario.objects.get(pk=origem_pk).origem_recurso
        except RecursoOrcamentario.DoesNotExist:
            origem = post_data.get('destino_origem_recurso', '').strip()
    else:
        origem = post_data.get('destino_origem_recurso', '').strip()

    if not (setor_ref and origem and natureza and rubrica):
        return post_data  # campos insuficientes — o form vai reclamar normalmente

    try:
        if setor_ref.isdigit():
            setor = Setor.objects.get(pk=setor_ref)
        else:
            setor = Setor.objects.get(sigla=setor_ref)
    except Setor.DoesNotExist:
        return post_data

    destino, _ = RecursoOrcamentario.objects.get_or_create(
        ano_fiscal=ano_atual,
        setor=setor,
        origem_recurso=origem,
        natureza=natureza,
        rubrica=rubrica,
        defaults={
            'valor_orcamentario': Decimal('0.00'),
            'criado_por': user,
        },
    )
    data = post_data.copy()
    data['destino'] = str(destino.pk)
    return data


@gestor_financeiro_required
def transferencia_create(request):
    ano_atual = timezone.now().year

    if request.method == 'POST':
        post_data = _resolver_destino(request.POST, request.user, ano_atual)
        form = TransferenciaForm(post_data, request.FILES or None)
    else:
        form = TransferenciaForm()

    if request.method == 'POST' and form.is_valid():
        t = form.save(commit=False)
        t.criada_por = request.user
        t.status = 'realizada'
        t.save()
        messages.success(request, 'Transferência registrada e efetivada.')
        return redirect('transferencia_list')

    ctx = _contexto_transferencia(ano_atual)
    ctx['form'] = form
    return render(request, 'orcamento/transferencia_form.html', ctx)

# ── Despesas ──────────────────────────────────────────────────────────────────

@gestor_financeiro_required
def despesa_list(request):
    qs = Despesa.objects.select_related('setor', 'recurso').order_by('-data_despesa')
    q        = request.GET.get('q', '')
    situacao = request.GET.get('situacao', '')
    natureza = request.GET.get('natureza', '')
    ano      = request.GET.get('ano', '')

    if q:
        qs = qs.filter(Q(discriminacao__icontains=q) | Q(nota_empenho__icontains=q))
    if situacao:
        qs = qs.filter(situacao=situacao)
    if natureza:
        qs = qs.filter(natureza=natureza)
    if ano:
        qs = qs.filter(data_despesa__year=ano)

    total_comprometido = qs.exclude(
        situacao__in=SituacaoDespesa.chaves_sem_impacto()
    ).aggregate(t=Sum('valor_comprometido'))['t'] or 0

    ctx = {
        'despesas':          qs,
        'total_comprometido': total_comprometido,
        'situacoes':         SituacaoDespesa.objects.all(),
        # .order_by() limpa o ordering padrão (-data_despesa) do Meta: sem isso o
        # DISTINCT considera a data completa e o mesmo ano aparece repetido no filtro.
        'anos': sorted(
            Despesa.objects.order_by().values_list('data_despesa__year', flat=True).distinct()
        ),
    }
    return render(request, 'orcamento/despesa_list.html', ctx)


def _is_rubrica_material(rubrica_str):
    s = (rubrica_str or '').lower()
    return 'consumo' in s or 'permanente' in s


def _contexto_despesa(ano_atual):
    def build_rubrica_meta():
        meta = {}
        for rb in Rubrica.objects.filter(ativo=True).select_related('natureza'):
            label = str(rb)
            natureza_key = _orc_norm(rb.natureza.nome if rb.natureza else '').lower()
            for alias in {label, rb.codigo, rb.nome}:
                key = _orc_norm(alias)
                if key:
                    meta[key] = {
                        'label': label,
                        'codigo': rb.codigo,
                        'natureza_key': natureza_key,
                    }
        return meta

    rubrica_meta = build_rubrica_meta()

    def rubrica_info(value):
        label = rubrica_normalizada(value)
        info = rubrica_meta.get(_orc_norm(label))
        if info:
            return info
        codigo = label.split(' - ', 1)[0].strip() if label else ''
        return {'label': label, 'codigo': codigo, 'natureza_key': ''}

    """Monta o contexto de recursos orçamentários para o formulário de despesa."""
    # Todos os anos: permite lançar despesas de exercícios anteriores. O ano é
    # escolhido na cascata do formulário (campo Ano Fiscal), evitando ambiguidade
    # entre recursos de mesma combinação setor/origem/natureza/rubrica em anos distintos.
    recursos = (
        RecursoOrcamentario.objects
        .select_related('setor')
        .order_by('-ano_fiscal', 'setor__sigla', 'origem_recurso', 'natureza', 'rubrica')
    )
    recursos_json = json.dumps([
        {
            'id': r.pk,
            'ano_fiscal': r.ano_fiscal,
            'setor': str(r.setor_id) if r.setor_id else '',
            'setor_label': _setor_catalog_label(r.setor),
            'origem_recurso': r.origem_recurso,
            'natureza': r.natureza,
            'natureza_key': _orc_norm(r.natureza).lower(),
            'rubrica': r.rubrica,
            'rubrica_norm': _orc_norm(rubrica_info(r.rubrica)['label']),
            'rubrica_codigo': rubrica_info(r.rubrica)['codigo'],
            'is_material': _is_rubrica_material(r.rubrica),
            'saldo_atual': float(r.saldo_atual),
        }
        for r in recursos
    ])
    itens = (
        RegistroPrecoItem.objects
        .select_related('registro', 'registro__pregao')
        .order_by('registro__numero_pregao', 'numero_item')
    )
    licitacao_itens = []
    for item in itens:
        rub_info = rubrica_info(item.rubrica)
        situacao = item.situacao_vigencia
        licitacao_itens.append({
            'id': item.pk,
            'pregao': item.registro.numero_pregao,
            'origem': item.registro.get_origem_display(),
            'origem_key': item.registro.origem,
            'item': item.numero_item,
            'material_licitado': item.material_licitado,
            'material': item.material,
            'rubrica': rub_info['label'] or item.rubrica,
            'rubrica_norm': _orc_norm(rub_info['label'] or item.rubrica),
            'rubrica_codigo': rub_info['codigo'],
            'natureza_key': rub_info['natureza_key'],
            'unidade': item.unidade,
            'valor': float(item.valor or 0),
            'saldo_unidade': item.saldo_unidade,
            'validade': item.validade_efetiva.isoformat() if item.validade_efetiva else '',
            'situacao': situacao,
            'situacao_label': _registro_situacao_label(situacao),
            'pregao_id': item.registro.pregao_id,
            'marca': item.marca,
        })
    pdi_hier = []
    for p in PdiPerspectiva.objects.prefetch_related('objetivos__indicadores').order_by('ordem', 'nome'):
        pdi_hier.append({
            'id': p.pk,
            'nome': p.nome,
            'codigo': p.codigo or '',
            'objetivos': [
                {
                    'id': o.pk,
                    'nome': o.nome,
                    'codigo': o.codigo or '',
                    'perspectiva_id': p.pk,
                    'indicadores': [
                        {'id': i.pk, 'nome': i.nome, 'codigo': i.codigo or '', 'objetivo_id': o.pk}
                        for i in o.indicadores.all()
                    ],
                }
                for o in p.objetivos.all()
            ],
        })

    return {
        'recursos_json': recursos_json,
        'licitacao_itens_json': json.dumps(licitacao_itens),
        'pdi_json': json.dumps(pdi_hier),
        'ano_atual': ano_atual,
    }


@gestor_financeiro_required
def despesa_create(request):
    from datetime import date
    ano_atual = date.today().year
    pregao_pk = request.GET.get('pregao')
    initial = {'pregao': pregao_pk} if pregao_pk else {}
    registro_item_pk = request.GET.get('registro_preco_item')
    if registro_item_pk:
        registro_item = get_object_or_404(
            RegistroPrecoItem.objects.select_related('registro'),
            pk=registro_item_pk,
        )
        initial.update({
            'registro_preco_item': registro_item.pk,
            'pregao': registro_item.registro.pregao_id,
            'pregao_ref': registro_item.registro.numero_pregao,
            'discriminacao': registro_item.material,
            'valor_unitario': registro_item.valor,
            'valor_comprometido': registro_item.valor,
            'quantidade': 1,
            'categoria_material': registro_item.marca,
        })
    itens_json = request.POST.get('despesa_itens_json', '').strip()
    if request.method == 'POST' and itens_json:
        form = DespesaForm(request.POST, initial=initial)
        erros = []
        try:
            itens_payload = json.loads(itens_json)
        except json.JSONDecodeError:
            itens_payload = []
            erros.append('A lista de itens da despesa não pôde ser lida.')

        recurso = None
        recurso_id = request.POST.get('recurso') or request.POST.get('id_recurso')
        if recurso_id:
            recurso = RecursoOrcamentario.objects.select_related('setor').filter(pk=recurso_id).first()
        if not recurso:
            erros.append('Selecione o recurso que será utilizado para esta despesa.')

        data_despesa = _registro_parse_date(request.POST.get('data_despesa'))
        if not data_despesa:
            erros.append('Informe a data da despesa.')

        situacao = request.POST.get('situacao') or 'empenhada'
        situacoes_validas = set(SituacaoDespesa.chaves_ativas())
        if situacao not in situacoes_validas:
            erros.append('Informe uma situação válida para a despesa.')

        if not itens_payload:
            erros.append('Adicione pelo menos um item à despesa.')

        item_ids = [
            row.get('registro_preco_item') for row in itens_payload
            if row.get('registro_preco_item') and not row.get('is_livre')
        ]
        item_map = {
            str(item.pk): item
            for item in RegistroPrecoItem.objects.select_related('registro', 'registro__pregao').filter(pk__in=item_ids)
        }

        linhas = []
        total_lote = Decimal('0')
        recurso_rubrica = rubrica_normalizada(recurso.rubrica) if recurso else ''
        recurso_codigo = recurso_rubrica.split(' - ', 1)[0].strip() if recurso_rubrica else ''
        for idx, row in enumerate(itens_payload, start=1):
            quantidade = _orc_parse_decimal(row.get('quantidade')) or Decimal('0')
            valor_unitario = _orc_parse_decimal(row.get('valor_unitario')) or Decimal('0')
            pdi_links = row.get('pdi_links', [])

            if row.get('is_livre'):
                discriminacao_livre = str(row.get('discriminacao') or '').strip()[:500]
                if not discriminacao_livre:
                    erros.append(f'Linha {idx}: informe a descrição do bem ou serviço.')
                    continue
                if quantidade <= 0:
                    erros.append(f'Linha {idx}: informe quantidade maior que zero.')
                    continue
                if valor_unitario <= 0:
                    erros.append(f'Linha {idx}: informe valor unitário maior que zero.')
                    continue
                valor_comprometido = (quantidade * valor_unitario).quantize(Decimal('0.01'))
                total_lote += valor_comprometido
                linhas.append({'is_livre': True, 'item': None,
                               'discriminacao': discriminacao_livre,
                               'quantidade': quantidade, 'valor_unitario': valor_unitario,
                               'valor_comprometido': valor_comprometido, 'pdi_links': pdi_links})
                continue

            item = item_map.get(str(row.get('registro_preco_item')))
            if not item:
                erros.append(f'Linha {idx}: item de licitação não encontrado.')
                continue
            item_rubrica = rubrica_normalizada(item.rubrica)
            item_codigo = item_rubrica.split(' - ', 1)[0].strip() if item_rubrica else ''
            rubrica_compativel = (
                _orc_norm(item_rubrica) == _orc_norm(recurso_rubrica)
                or (item_codigo and recurso_codigo and item_codigo == recurso_codigo)
            )
            if recurso and not rubrica_compativel:
                erros.append(f'Linha {idx}: a rubrica do item não corresponde à rubrica do recurso selecionado.')
                continue
            # Itens de licitação vencidos são permitidos: a equipe precisa registrar
            # despesas vinculadas a pregões cuja validade já expirou.
            if quantidade <= 0:
                erros.append(f'Linha {idx}: informe quantidade maior que zero.')
                continue
            if valor_unitario <= 0:
                erros.append(f'Linha {idx}: informe valor unitário maior que zero.')
                continue
            valor_comprometido = (quantidade * valor_unitario).quantize(Decimal('0.01'))
            total_lote += valor_comprometido
            linhas.append({'is_livre': False, 'item': item, 'discriminacao': None,
                           'quantidade': quantidade, 'valor_unitario': valor_unitario,
                           'valor_comprometido': valor_comprometido, 'pdi_links': pdi_links})

        if recurso and total_lote > recurso.saldo_atual:
            erros.append(f'Valor total excede o saldo disponível do recurso selecionado (R$ {recurso.saldo_atual:,.2f}).')

        if erros:
            for erro in erros[:5]:
                messages.error(request, erro)
            if len(erros) > 5:
                messages.error(request, f'Há mais {len(erros) - 5} pendência(s) na despesa.')
            ctx = {'form': form}
            ctx.update(_contexto_despesa(ano_atual))
            return render(request, 'orcamento/despesa_form.html', ctx)

        def optional_fk(model, value):
            return model.objects.filter(pk=value).first() if str(value or '').isdigit() else None

        with transaction.atomic():
            for linha in linhas:
                pdi_links = linha['pdi_links']
                perspectiva = None
                objetivo = None
                indicador = None
                if pdi_links:
                    first_link = pdi_links[0]
                    if first_link.get('perspectiva_id'):
                        perspectiva = optional_fk(PdiPerspectiva, first_link['perspectiva_id'])
                    if first_link.get('objetivo_id'):
                        objetivo = optional_fk(PdiObjetivoEstrategico, first_link['objetivo_id'])
                    if first_link.get('indicador_id'):
                        indicador = optional_fk(PdiIndicador, first_link['indicador_id'])
                common = dict(
                    data_despesa=data_despesa,
                    requisicao=request.POST.get('requisicao', ''),
                    nota_empenho=request.POST.get('nota_empenho', ''),
                    quantidade=linha['quantidade'],
                    valor_unitario=linha['valor_unitario'],
                    valor_comprometido=linha['valor_comprometido'],
                    rubrica=recurso.rubrica,
                    setor=recurso.setor,
                    recurso=recurso,
                    situacao=situacao,
                    observacao=request.POST.get('observacao', ''),
                    natureza=recurso.natureza,
                    perspectiva_pdi=perspectiva,
                    objetivo_pdi=objetivo,
                    indicador_pdi=indicador,
                    pdi_vinculos=pdi_links,
                    criada_por=request.user,
                )
                if linha['is_livre']:
                    Despesa.objects.create(
                        **common,
                        pregao_ref='', pregao=None, registro_preco_item=None,
                        discriminacao=linha['discriminacao'],
                        categoria_material='',
                    )
                else:
                    item = linha['item']
                    Despesa.objects.create(
                        **common,
                        pregao_ref=item.registro.numero_pregao,
                        pregao=item.registro.pregao,
                        registro_preco_item=item,
                        discriminacao=item.material,
                        categoria_material=item.marca,
                    )
        messages.success(request, f'{len(linhas)} item(ns) de despesa registrado(s).')
        return redirect('despesa_list')

    form = DespesaForm(request.POST or None, initial=initial)
    if form.is_valid():
        d = form.save(commit=False)
        d.criada_por = request.user
        d.save()
        messages.success(request, 'Despesa registrada.')
        return redirect('despesa_list')
    ctx = {'form': form}
    ctx.update(_contexto_despesa(ano_atual))
    return render(request, 'orcamento/despesa_form.html', ctx)


@gestor_financeiro_required
def despesa_edit(request, pk):
    from datetime import date
    from decimal import ROUND_HALF_UP
    obj = get_object_or_404(Despesa, pk=pk)
    ano_atual = obj.recurso.ano_fiscal if obj.recurso_id else date.today().year
    form = DespesaForm(request.POST or None, instance=obj)
    if form.is_valid():
        despesa = form.save(commit=False)
        # Persiste itens da nova UI (qtd, val_unit, pdi_vinculos) a partir do JSON
        try:
            itens_payload = json.loads(request.POST.get('despesa_itens_json') or '[]')
        except (ValueError, TypeError):
            itens_payload = []
        if itens_payload:
            primeiro = itens_payload[0] or {}
            try:
                qtd = Decimal(str(primeiro.get('quantidade') or '1').replace(',', '.'))
                if qtd <= 0:
                    qtd = Decimal('1')
            except (InvalidOperation, ValueError):
                qtd = Decimal('1')
            try:
                vu = Decimal(str(primeiro.get('valor_unitario') or '0').replace(',', '.'))
                if vu < 0:
                    vu = Decimal('0')
            except (InvalidOperation, ValueError):
                vu = Decimal('0')
            despesa.quantidade = qtd
            despesa.valor_unitario = vu
            despesa.valor_comprometido = (qtd * vu).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            if primeiro.get('is_livre') and primeiro.get('discriminacao'):
                despesa.discriminacao = str(primeiro.get('discriminacao'))[:1000]
            try:
                despesa.pdi_vinculos = primeiro.get('pdi_links') or []
            except Exception:
                pass
        despesa.save()
        messages.success(request, 'Despesa atualizada.')
        return redirect('despesa_list')

    itens_iniciais_json = '[]'
    if not request.POST:
        try:
            pdi_links = list(obj.pdi_vinculos or [])
        except Exception:
            pdi_links = []
        if not pdi_links and (obj.perspectiva_pdi_id or obj.objetivo_pdi_id or obj.indicador_pdi_id):
            pdi_links = [{
                'perspectiva_id':   obj.perspectiva_pdi_id,
                'perspectiva_nome': obj.perspectiva_pdi.nome if obj.perspectiva_pdi_id else '',
                'objetivo_id':      obj.objetivo_pdi_id,
                'objetivo_nome':    obj.objetivo_pdi.nome if obj.objetivo_pdi_id else '',
                'indicador_id':     obj.indicador_pdi_id,
                'indicador_nome':   obj.indicador_pdi.nome if obj.indicador_pdi_id else '',
            }]

        qtd = obj.quantidade or Decimal('1')
        if qtd <= 0:
            qtd = Decimal('1')
        val_comp = obj.valor_comprometido or Decimal('0')
        # Derivar val_unit de valor_comprometido (fonte da verdade financeira)
        if val_comp > 0:
            val_unit = (val_comp / qtd).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            val_unit = obj.valor_unitario or Decimal('0')
        # Quantidade como inteiro quando for inteiro
        qtd_str = str(int(qtd)) if qtd == int(qtd) else str(qtd.normalize())
        val_unit_str = format(val_unit, 'f')

        if _is_rubrica_material(obj.rubrica) and obj.registro_preco_item_id:
            raw = json.dumps([{
                'registro_preco_item': str(obj.registro_preco_item_id),
                'quantidade': qtd_str,
                'valor_unitario': val_unit_str,
                'saldo_unidade': None,
                'unidade': '',
                'pdi_links': pdi_links,
            }], ensure_ascii=False)
        else:
            raw = json.dumps([{
                'id': 'livre_1',
                'discriminacao': obj.discriminacao or '',
                'rubrica': obj.rubrica or '',
                'quantidade': qtd_str,
                'valor_unitario': val_unit_str,
                'pdi_links': pdi_links,
                'is_livre': True,
            }], ensure_ascii=False)
        itens_iniciais_json = raw.replace('</', '<\\/')

    ctx = {'form': form, 'object': obj, 'itens_iniciais_json': itens_iniciais_json}
    ctx.update(_contexto_despesa(ano_atual))
    return render(request, 'orcamento/despesa_form.html', ctx)


@gestor_financeiro_required
def despesa_delete(request, pk):
    obj = get_object_or_404(Despesa, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Despesa excluída.')
        return redirect('despesa_list')
    return render(request, 'orcamento/despesa_confirm_delete.html', {'object': obj})


@gestor_financeiro_required
def despesa_excluir_lote(request):
    if request.method != 'POST':
        return redirect('despesa_list')
    pks = request.POST.getlist('pks')
    if not pks:
        messages.warning(request, 'Nenhuma despesa selecionada.')
        return redirect('despesa_list')
    excluidos = Despesa.objects.filter(pk__in=pks).count()
    Despesa.objects.filter(pk__in=pks).delete()
    messages.success(request, f'{excluidos} despesa(s) excluída(s).')
    return redirect('despesa_list')


# ── Administração — Usuários ─────────────────────────────────────────────────

# PERFIL_POR_TIPO_SETOR expandido para incluir os perfis financeiros
_PERFIL_POR_TIPO_ORC = {
    tipo: perfis + ['gestor_financeiro', 'ordenador_despesa']
    for tipo, perfis in {
        'centro':         ['admin', 'gestor_irp', 'diretor_centro'],
        'direcao':        ['admin', 'gestor_irp', 'diretor_centro'],
        'departamento':   ['admin', 'gestor_irp', 'aprovador_setor'],
        'coordenacao_g':  ['admin', 'gestor_irp', 'aprovador_setor', 'respondente'],
        'coordenacao_pg': ['admin', 'gestor_irp', 'aprovador_setor', 'respondente'],
        'administrativo': ['admin', 'gestor_irp', 'aprovador_setor', 'respondente'],
        'laboratorio':    ['admin', 'gestor_irp', 'respondente'],
        'secretaria':     ['admin', 'gestor_irp', 'respondente'],
    }.items()
}


@gestor_financeiro_required
def orc_usuario_list(request):
    from django.contrib.auth.models import User as AuthUser
    usuarios = AuthUser.objects.select_related('perfil').order_by(
        'perfil__nome_completo', 'username'
    )
    return render(request, 'orcamento/admin/usuario_list.html', {'usuarios': usuarios})


@gestor_financeiro_required
def orc_usuario_create(request):
    from core.views import UsuarioForm, TIPO_SETOR_CHOICES
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.encoding import force_bytes
    from django.utils.http import urlsafe_base64_encode
    from django.core.mail import send_mail
    from django.db import transaction as db_transaction

    todos_setores = Setor.objects.filter(ativo=True).order_by('nome')
    if request.method == 'POST':
        form = UsuarioForm(request.POST, request.FILES)
        if form.is_valid():
            with db_transaction.atomic():
                user = form.save()
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                token = default_token_generator.make_token(user)
                link = request.build_absolute_uri(f'/ativar-conta/{uid}/{token}/')
                try:
                    send_mail(
                        subject='Ative sua conta – IRP CT/UFPB',
                        message=(
                            f'Olá, {user.username},\n\n'
                            'Sua conta foi criada. Acesse o link abaixo para definir sua senha '
                            f'(válido por 24h):\n\n{link}\n\nAtenciosamente,\nSistema CT/UFPB'
                        ),
                        from_email=None,
                        recipient_list=[user.email],
                        fail_silently=False,
                    )
                    messages.success(request, f'Usuário "{user.username}" criado. E-mail de ativação enviado.')
                except Exception as e:
                    messages.warning(request, f'Usuário "{user.username}" criado, mas erro ao enviar e-mail: {e}')
            return redirect('orc_usuario_list')
    else:
        form = UsuarioForm()
    return render(request, 'orcamento/admin/usuario_form.html', {
        'form': form,
        'titulo_pagina': 'Novo Usuário',
        'todos_setores': todos_setores,
        'tipo_setor_choices': TIPO_SETOR_CHOICES,
        'perfil_por_tipo_json': json.dumps(_PERFIL_POR_TIPO_ORC),
    })


@gestor_financeiro_required
def orc_usuario_edit(request, pk):
    from django.contrib.auth.models import User as AuthUser
    from core.views import UsuarioForm, TIPO_SETOR_CHOICES
    from core.models import PerfilUsuario

    usuario = get_object_or_404(AuthUser, pk=pk)
    try:
        perfil = usuario.perfil
    except PerfilUsuario.DoesNotExist:
        perfil = None

    todos_setores = Setor.objects.filter(ativo=True).order_by('nome')
    if request.method == 'POST':
        form = UsuarioForm(request.POST, request.FILES, instance=usuario, perfil=perfil)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário atualizado.')
            return redirect('orc_usuario_list')
    else:
        form = UsuarioForm(instance=usuario, perfil=perfil)

    return render(request, 'orcamento/admin/usuario_form.html', {
        'form': form,
        'usuario': usuario,
        'titulo_pagina': 'Editar Usuário',
        'todos_setores': todos_setores,
        'tipo_setor_choices': TIPO_SETOR_CHOICES,
        'perfil': perfil,
        'perfil_por_tipo_json': json.dumps(_PERFIL_POR_TIPO_ORC),
    })


@gestor_financeiro_required
def orc_usuario_toggle(request, pk):
    from django.contrib.auth.models import User as AuthUser
    usuario = get_object_or_404(AuthUser, pk=pk)
    if request.method == 'POST' and usuario != request.user:
        usuario.is_active = not usuario.is_active
        usuario.save()
        estado = 'ativado' if usuario.is_active else 'desativado'
        messages.success(request, f'Usuário {estado}.')
    return redirect('orc_usuario_list')


@gestor_financeiro_required
def orc_usuario_apagar(request, pk):
    from django.contrib.auth.models import User as AuthUser
    usuario = get_object_or_404(AuthUser, pk=pk)
    if request.method == 'POST':
        if usuario == request.user:
            messages.error(request, 'Você não pode apagar sua própria conta.')
            return redirect('orc_usuario_list')
        nome = getattr(getattr(usuario, 'perfil', None), 'nome_completo', None) or usuario.username
        usuario.delete()
        messages.success(request, f'Usuário "{nome}" apagado.')
    return redirect('orc_usuario_list')


@gestor_financeiro_required
def orc_cadastros(request):
    naturezas = NaturezaRecurso.objects.all()
    rubricas  = Rubrica.objects.select_related('natureza').all()
    origens   = OrigemRecurso.objects.all()
    situacoes = SituacaoDespesa.objects.all()
    return render(request, 'orcamento/admin/cadastros.html', {
        'naturezas':         naturezas,
        'rubricas':          rubricas,
        'naturezas_choices': naturezas,
        'origens':           origens,
        'situacoes':         situacoes,
    })


# ── Cadastros — Natureza do Recurso ──────────────────────────────────────────

@gestor_financeiro_required
def orc_natureza_create(request):
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if nome:
            if NaturezaRecurso.objects.filter(nome__iexact=nome).exists():
                messages.warning(request, f'Já existe uma natureza com o nome "{nome}".')
            else:
                ordem = NaturezaRecurso.objects.count() + 1
                NaturezaRecurso.objects.create(nome=nome, ordem=ordem)
                messages.success(request, f'Natureza "{nome}" criada com sucesso.')
        else:
            messages.error(request, 'Informe um nome para a natureza.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_natureza_edit(request, pk):
    obj = get_object_or_404(NaturezaRecurso, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if nome:
            if NaturezaRecurso.objects.filter(nome__iexact=nome).exclude(pk=pk).exists():
                messages.warning(request, f'Já existe outra natureza com o nome "{nome}".')
            else:
                obj.nome = nome
                obj.save()
                messages.success(request, f'Natureza atualizada para "{nome}".')
        else:
            messages.error(request, 'Informe um nome para a natureza.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_natureza_toggle(request, pk):
    obj = get_object_or_404(NaturezaRecurso, pk=pk)
    if request.method == 'POST':
        obj.ativo = not obj.ativo
        obj.save()
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_natureza_delete(request, pk):
    obj = get_object_or_404(NaturezaRecurso, pk=pk)
    if request.method == 'POST':
        nome = obj.nome
        obj.delete()
        messages.success(request, f'Natureza "{nome}" excluída.')
    return redirect('orc_cadastros')


# ── Cadastros — Rubrica ───────────────────────────────────────────────────────

@gestor_financeiro_required
def orc_rubrica_create(request):
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        nome   = request.POST.get('nome', '').strip()
        nat_pk = request.POST.get('natureza', '').strip()
        if not codigo or not nome:
            messages.error(request, 'Preencha o código e o nome da rubrica.')
        else:
            natureza = NaturezaRecurso.objects.filter(pk=nat_pk).first() if nat_pk else None
            ordem    = Rubrica.objects.count() + 1
            Rubrica.objects.create(codigo=codigo, nome=nome, natureza=natureza, ordem=ordem)
            messages.success(request, f'Rubrica "{codigo} - {nome}" criada com sucesso.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_rubrica_edit(request, pk):
    obj = get_object_or_404(Rubrica, pk=pk)
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        nome   = request.POST.get('nome', '').strip()
        nat_pk = request.POST.get('natureza', '').strip()
        if not codigo or not nome:
            messages.error(request, 'Preencha o código e o nome da rubrica.')
        else:
            obj.codigo   = codigo
            obj.nome     = nome
            obj.natureza = NaturezaRecurso.objects.filter(pk=nat_pk).first() if nat_pk else None
            obj.save()
            messages.success(request, f'Rubrica "{codigo} - {nome}" atualizada.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_rubrica_toggle(request, pk):
    obj = get_object_or_404(Rubrica, pk=pk)
    if request.method == 'POST':
        obj.ativo = not obj.ativo
        obj.save()
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_rubrica_delete(request, pk):
    obj = get_object_or_404(Rubrica, pk=pk)
    if request.method == 'POST':
        label = str(obj)
        obj.delete()
        messages.success(request, f'Rubrica "{label}" excluída.')
    return redirect('orc_cadastros')


# ── Cadastros — Origem do Recurso ─────────────────────────────────────────────

@gestor_financeiro_required
def orc_origem_create(request):
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if nome:
            if OrigemRecurso.objects.filter(nome__iexact=nome).exists():
                messages.warning(request, f'Já existe uma origem com o nome "{nome}".')
            else:
                ordem = OrigemRecurso.objects.count() + 1
                OrigemRecurso.objects.create(nome=nome, ordem=ordem)
                messages.success(request, f'Origem "{nome}" criada com sucesso.')
        else:
            messages.error(request, 'Informe um nome para a origem.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_origem_edit(request, pk):
    obj = get_object_or_404(OrigemRecurso, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if nome:
            if OrigemRecurso.objects.filter(nome__iexact=nome).exclude(pk=pk).exists():
                messages.warning(request, f'Já existe outra origem com o nome "{nome}".')
            else:
                obj.nome = nome
                obj.save()
                messages.success(request, f'Origem atualizada para "{nome}".')
        else:
            messages.error(request, 'Informe um nome para a origem.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_origem_toggle(request, pk):
    obj = get_object_or_404(OrigemRecurso, pk=pk)
    if request.method == 'POST':
        obj.ativo = not obj.ativo
        obj.save()
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_origem_delete(request, pk):
    obj = get_object_or_404(OrigemRecurso, pk=pk)
    if request.method == 'POST':
        nome = obj.nome
        obj.delete()
        messages.success(request, f'Origem "{nome}" excluída.')
    return redirect('orc_cadastros')


# ── Licitações do IRP ─────────────────────────────────────────────────────────

@gestor_financeiro_required
def orc_situacao_despesa_create(request):
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        impacta_saldo = request.POST.get('impacta_saldo') == 'on'
        badge = request.POST.get('badge', '').strip() or 'bg-secondary'
        if nome:
            if SituacaoDespesa.objects.filter(nome__iexact=nome).exists():
                messages.warning(request, f'Já existe uma situação de despesa com o nome "{nome}".')
            else:
                ordem = SituacaoDespesa.objects.count() + 1
                SituacaoDespesa.objects.create(
                    nome=nome,
                    ordem=ordem,
                    impacta_saldo=impacta_saldo,
                    badge=badge,
                )
                messages.success(request, f'Situação "{nome}" criada com sucesso.')
        else:
            messages.error(request, 'Informe um nome para a situação da despesa.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_situacao_despesa_edit(request, pk):
    obj = get_object_or_404(SituacaoDespesa, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if nome:
            if SituacaoDespesa.objects.filter(nome__iexact=nome).exclude(pk=pk).exists():
                messages.warning(request, f'Já existe outra situação de despesa com o nome "{nome}".')
            else:
                obj.nome = nome
                obj.impacta_saldo = request.POST.get('impacta_saldo') == 'on'
                obj.badge = request.POST.get('badge', '').strip() or 'bg-secondary'
                obj.save()
                messages.success(request, f'Situação atualizada para "{nome}".')
        else:
            messages.error(request, 'Informe um nome para a situação da despesa.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_situacao_despesa_toggle(request, pk):
    obj = get_object_or_404(SituacaoDespesa, pk=pk)
    if request.method == 'POST':
        obj.ativo = not obj.ativo
        obj.save()
    return redirect('orc_cadastros')


@gestor_financeiro_required
def orc_situacao_despesa_delete(request, pk):
    obj = get_object_or_404(SituacaoDespesa, pk=pk)
    if request.method == 'POST':
        nome = obj.nome
        if Despesa.objects.filter(situacao=obj.chave).exists():
            messages.error(
                request,
                f'Não é possível excluir "{nome}" porque existem despesas com essa situação. Desative a categoria se não quiser mais usá-la.',
            )
        else:
            obj.delete()
            messages.success(request, f'Situação "{nome}" excluída.')
    return redirect('orc_cadastros')


@gestor_financeiro_required
def licitacao_list(request):
    pregoes = Pregao.objects.select_related('irp').prefetch_related(
        'itens_pregao__item'
    ).order_by('-atualizado_em')
    return render(request, 'orcamento/licitacao_list.html', {'pregoes': pregoes})


REGISTRO_PRECO_HEADERS = [
    'Pregão', 'Material Licitado', 'Item', 'Material', 'Rubrica', 'Unidade', 'Qtd. Emp.',
    'Marca', 'Valor', 'Saldo UFPB', 'Saldo Unidade', 'Validade',
]


def _registro_parse_date(value):
    from datetime import date, datetime
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value).date()
        except Exception:
            return None

    text = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _registro_preco_chave(*parts):
    raw = '|'.join(_orc_norm(p) for p in parts)
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()


def _registro_situacao_label(situacao):
    return {
        'vigente': 'Vigente',
        'a_vencer': 'A vencer',
        'vencido': 'Vencido',
        'sem_validade': 'Sem validade',
    }.get(situacao, situacao)


def _add_um_ano(data):
    if not data:
        return None
    try:
        return data.replace(year=data.year + 1)
    except ValueError:
        return data.replace(year=data.year + 1, day=28)


def _registro_validade_irp(pregao):
    data_base = getattr(pregao, 'data_publicacao', None) or pregao.data_homologacao
    return _add_um_ano(data_base)


def _format_quantidade_saldo_unidade(value):
    if value in (None, ''):
        return ''
    quant = Decimal(value)
    text = f'{quant:.3f}'.rstrip('0').rstrip('.')
    return text.replace('.', ',')


def _normalizar_unidade_registro(value):
    text = str(value or '').strip()
    if not text:
        return ''
    if _orc_norm(text).replace('.', '') in {'UNIDADE', 'UNID', 'UND'}:
        return 'UND'
    return text


def _sync_registro_preco_irp(user=None):
    criados = 0
    atualizados = 0
    pregoes = (
        Pregao.objects.filter(status='homologado')
        .select_related('irp')
        .prefetch_related('itens_pregao__item')
    )

    for pregao in pregoes:
        registro = RegistroPrecoVigente.objects.filter(origem='rp', pregao=pregao).first()
        if not registro:
            registro = RegistroPrecoVigente(origem='rp', pregao=pregao, criado_por=user)

        validade = _registro_validade_irp(pregao)
        registro.numero_pregao = pregao.numero or f'IRP-{pregao.pk}'
        registro.objeto = str(pregao.irp)
        registro.data_homologacao = pregao.data_homologacao
        registro.validade = validade
        registro.save()

        qty_homologada = {
            row['item_id']: row['total'] or Decimal('0')
            for row in HomologacaoSetorItem.objects
            .filter(
                homologacao__irp=pregao.irp,
                homologacao__status='homologada',
                quantidade_aprovada__isnull=False,
            )
            .values('item_id')
            .annotate(total=Sum('quantidade_aprovada'))
        }
        itens = pregao.itens_pregao.filter(situacao='licitado').select_related('item')
        for pi in itens:
            item = pi.item
            chave = f'rp:{pi.pk}'
            valor = pi.preco_licitado or item.preco_estimado
            saldo_unidade = _format_quantidade_saldo_unidade(
                qty_homologada.get(item.pk, Decimal('0'))
            )
            _, created = RegistroPrecoItem.objects.update_or_create(
                chave_importacao=chave,
                defaults={
                    'registro': registro,
                    'pregao_item': pi,
                    'numero_item': str(item.numero),
                    'material_licitado': (getattr(pregao.irp, 'descricao', '') or '').strip(),
                    'material': item.descricao,
                    'rubrica': rubrica_normalizada(item.rubrica),
                    'unidade': _normalizar_unidade_registro(item.unidade),
                    'quantidade_empenhada': item.quantidade_total,
                    'marca': '',
                    'valor': valor,
                    'saldo': None,
                    'saldo_unidade': saldo_unidade,
                    'validade': validade,
                }
            )
            criados += 1 if created else 0
            atualizados += 0 if created else 1

    return criados, atualizados


@gestor_financeiro_required
def registro_preco_list(request):
    q = request.GET.get('q', '').strip()
    origem = request.GET.get('origem', '').strip()
    situacao = request.GET.get('situacao', '').strip()

    qs = (
        RegistroPrecoItem.objects
        .select_related('registro', 'pregao_item__item')
        .order_by('registro__numero_pregao', 'numero_item')
    )
    if q:
        qs = qs.filter(
            Q(material_licitado__icontains=q) |
            Q(material__icontains=q) |
            Q(marca__icontains=q) |
            Q(rubrica__icontains=q) |
            Q(registro__numero_pregao__icontains=q) |
            Q(numero_item__icontains=q)
        )
    if origem:
        qs = qs.filter(registro__origem=origem)

    itens = list(qs)
    if situacao:
        itens = [i for i in itens if i.situacao_vigencia == situacao]

    registros = {i.registro_id: i.registro for i in itens}
    stats = {
        'total': len(itens),
        'vigentes': sum(1 for i in itens if i.situacao_vigencia == 'vigente'),
        'a_vencer': sum(1 for i in itens if i.situacao_vigencia == 'a_vencer'),
        'vencidos': sum(1 for i in itens if i.situacao_vigencia == 'vencido'),
        'pregoes_total': len(registros),
        'pregoes_vigentes': sum(1 for r in registros.values() if r.situacao_vigencia == 'vigente'),
    }
    ctx = {
        'itens': itens,
        'stats': stats,
        'filtros': {'q': q, 'origem': origem, 'situacao': situacao},
        'situacao_label': _registro_situacao_label,
    }
    return render(request, 'orcamento/registro_preco_list.html', ctx)


@gestor_financeiro_required
def registro_preco_sync_irp(request):
    if request.method != 'POST':
        return redirect('registro_preco_list')
    criados, atualizados = _sync_registro_preco_irp(request.user)
    messages.success(
        request,
        f'Registros do módulo IRP sincronizados: {criados} criado(s), {atualizados} atualizado(s).'
    )
    return redirect('registro_preco_list')


@gestor_financeiro_required
def registro_preco_excluir_lote(request):
    if request.method != 'POST':
        return redirect('registro_preco_list')
    pks = request.POST.getlist('pks')
    if not pks:
        messages.info(request, 'Nenhum item selecionado para exclusão.')
        return redirect('registro_preco_list')
    qs = RegistroPrecoItem.objects.filter(pk__in=pks)
    total = qs.count()
    qs.delete()
    messages.success(request, f'{total} item(ns) de registro de preço excluído(s).')
    return redirect('registro_preco_list')


@gestor_financeiro_required
def registro_preco_template_xlsx(request):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registro de Preço'

    header_fill = PatternFill('solid', fgColor='B7DEE8')
    border = Border(bottom=Side(style='thin', color='9FBAC4'))
    for col, header in enumerate(REGISTRO_PRECO_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='000000')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    widths = [18, 60, 10, 60, 38, 16, 14, 24, 16, 16, 16, 16]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    ws.freeze_panes = 'A2'

    ws_ref = wb.create_sheet('Referências')
    rubricas = [str(r) for r in Rubrica.objects.filter(ativo=True).order_by('ordem', 'codigo')]
    ws_ref['A1'] = 'Rubrica'
    ws_ref['A1'].font = Font(bold=True)
    for row, value in enumerate(rubricas, 2):
        ws_ref.cell(row=row, column=1, value=value)
    ws_ref.column_dimensions['A'].width = 42

    if rubricas:
        dv = DataValidation(type='list', formula1=f"'Referências'!$A$2:$A${len(rubricas) + 1}")
        dv.errorTitle = 'Rubrica inválida'
        dv.error = 'Escolha uma rubrica cadastrada na aba Referências.'
        dv.promptTitle = 'Rubrica'
        dv.prompt = 'Selecione uma rubrica cadastrada.'
        ws.add_data_validation(dv)
        dv.add('E2:E501')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_registro_preco_vigente.xlsx"'
    wb.save(response)
    return response


@gestor_financeiro_required
def registro_preco_importar(request):
    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            messages.error(request, 'Selecione uma planilha para importar.')
            return redirect('registro_preco_importar')

        try:
            rows = _orc_rows_from_xlsx(arquivo)
        except Exception as exc:
            messages.error(request, f'Não foi possível ler a planilha: {exc}')
            return redirect('registro_preco_importar')

        criados = atualizados = 0
        erros = []
        for idx, row in enumerate(rows, 2):
            numero_pregao = str(row.get(_orc_norm('Pregão')) or row.get(_orc_norm('Pregao')) or '').strip()
            numero_item = str(row.get(_orc_norm('Item')) or '').strip()
            material_licitado = str(row.get(_orc_norm('Material Licitado')) or '').strip()
            material = str(row.get(_orc_norm('Material')) or '').strip()
            rubrica_raw = str(row.get(_orc_norm('Rubrica')) or '').strip()
            unidade = _normalizar_unidade_registro(
                row.get(_orc_norm('Unidade')) or
                row.get(_orc_norm('Unid.')) or
                row.get(_orc_norm('Unid')) or
                row.get(_orc_norm('Und.')) or
                row.get(_orc_norm('Und')) or
                ''
            )
            marca = str(row.get(_orc_norm('Marca')) or '').strip()
            saldo_unidade = str(
                row.get(_orc_norm('Saldo Unidade')) or
                row.get(_orc_norm('Saldo Und.')) or
                row.get(_orc_norm('Saldo Und')) or
                ''
            ).strip()
            validade = _registro_parse_date(row.get(_orc_norm('Validade')))

            if not (numero_pregao and numero_item and material):
                erros.append(f'Linha {idx}: Pregão, Item e Material são obrigatórios.')
                continue
            try:
                rubrica = rubrica_normalizada(rubrica_raw, strict=bool(rubrica_raw))
            except ValueError as exc:
                erros.append(f'Linha {idx}: {exc}')
                continue

            quantidade_empenhada = _orc_parse_decimal(
                row.get(_orc_norm('Qtd. Emp.')) or row.get(_orc_norm('Qtd Emp'))
            )
            valor = _orc_parse_decimal(row.get(_orc_norm('Valor')))
            saldo = _orc_parse_decimal(
                row.get(_orc_norm('Saldo UFPB')) or
                row.get(_orc_norm('Saldo'))
            )
            chave = 'sipac:' + _registro_preco_chave(numero_pregao, numero_item, material, marca, validade or '')

            registro = RegistroPrecoVigente.objects.filter(
                origem='sipac',
                numero_pregao=numero_pregao,
            ).first()
            if not registro:
                registro = RegistroPrecoVigente(
                    origem='sipac',
                    numero_pregao=numero_pregao,
                    criado_por=request.user,
                )
            registro.objeto = registro.objeto or f'Pregão {numero_pregao}'
            if validade and (not registro.validade or validade > registro.validade):
                registro.validade = validade
            registro.save()

            _, created = RegistroPrecoItem.objects.update_or_create(
                chave_importacao=chave,
                defaults={
                    'registro': registro,
                    'pregao_item': None,
                    'numero_item': numero_item,
                    'material_licitado': material_licitado,
                    'material': material,
                    'rubrica': rubrica,
                    'unidade': unidade,
                    'quantidade_empenhada': quantidade_empenhada,
                    'marca': marca,
                    'valor': valor,
                    'saldo': saldo,
                    'saldo_unidade': saldo_unidade,
                    'validade': validade,
                }
            )
            criados += 1 if created else 0
            atualizados += 0 if created else 1

        if erros:
            messages.warning(request, 'Importação concluída com alertas: ' + ' | '.join(erros[:8]))
        messages.success(request, f'Importação SIPAC concluída: {criados} criado(s), {atualizados} atualizado(s).')
        return redirect('registro_preco_list')

    return render(request, 'orcamento/registro_preco_importar.html')
