"""Comando temporário para ler a planilha de setores e exibir os dados."""
import sys
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = 'Lê a planilha de setores e mostra os dados'

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stdout.write('Instalando openpyxl...')
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
            import openpyxl

        path = r'C:\Users\MarcelGois\Downloads\Hierarquia_CT_UFPB.xlsx'
        wb = openpyxl.load_workbook(path)
        self.stdout.write(f'Abas: {wb.sheetnames}')

        output_lines = [f'Abas: {wb.sheetnames}']
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            line = f'\n=== {sheet_name} ({ws.max_row} linhas x {ws.max_column} colunas) ==='
            self.stdout.write(line)
            output_lines.append(line)
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    row_str = str(row)
                    self.stdout.write(row_str)
                    output_lines.append(row_str)

        # Salva em arquivo para leitura posterior
        output_path = r'C:\Users\MarcelGois\Documents\Projeto - IRP App\xlsx_output.txt'
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(output_lines))
        self.stdout.write(f'\nSalvo em: {output_path}')
