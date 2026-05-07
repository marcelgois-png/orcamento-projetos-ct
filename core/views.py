import csv
import functools
import io
import json
import unicodedata
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode

from .forms import (IRPForm, ItemForm, ItemImportForm, PerfilForm,
                    SetorForm, UsuarioForm)
from .models import (IRP, Item, PerfilUsuario, Resposta, RespostaItem, Setor,
                      HomologacaoSetor, HomologacaoSetorItem,
                      Pregao, PregaoItem, PERFIL_POR_TIPO_SETOR, TIPO_SETOR_CHOICES)
from .rubricas import rubrica_catalog_labels, rubrica_normalizada


# ---------------------------------------------------------------------------
# HTMX: retorna subsetores filtrados por setor pai
# ---------------------------------------------------------------------------

def subsetores_ajax(request):
    """Retorna <option> HTML dos subsetores do setor informado (para HTMX).

    Aceita o ID do setor pai via parâmetro GET: setor_id ou setor_pai.
    """
    setor_id = (
        request.GET.get('setor_id') or
        request.GET.get('setor_pai') or
        request.GET.get('setor_pai_id') or
        ''
    )
    subsetores = []
    if setor_id:
        try:
            subsetores = list(
                Setor.objects.filter(pai_id=int(setor_id), ativo=True).order_by('nome')
            )
        except (ValueError, TypeError):
            pass
    html = '<option value="">--- Nenhum ---</option>'
    for s in subsetores:
        html += f'<option value="{s.pk}" data-tipo="{s.tipo}">{s.nome}</option>'
    return HttpResponse(html)


def _setor_pai_sub(setor):
    """Dado um Setor, retorna (pai, sub) onde pai é o setor raiz e sub é o filho
    (ou None se o próprio setor já for raiz)."""
    if setor is None:
        return None, None
    if setor.pai_id is None:
        return setor, None
    return setor.pai, setor


# ---------------------------------------------------------------------------
# Decorador de gestor
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Mapeamento setor.tipo → (grupo, cor) para gráficos
# ---------------------------------------------------------------------------
_NIVEL_SETOR = {
    'centro':         ('Setores Administrativos', '#283E4C'),
    'direcao':        ('Setores Administrativos', '#283E4C'),
    'administrativo': ('Setores Administrativos', '#283E4C'),
    'secretaria':     ('Setores Administrativos', '#283E4C'),
    'departamento':   ('Departamentos',           '#A5B1BB'),
    'laboratorio':    ('Departamentos',           '#A5B1BB'),
    'coordenacao_g':  ('Coordenações',            '#C6543C'),
    'coordenacao_pg': ('Coordenações',            '#C6543C'),
}


def _build_por_nivel_chart(respostas_qs):
    """Recebe um queryset de Resposta (já prefetchado) e retorna lista
    de dicts {label, nome, nivel, cor, valor} ordenada por valor desc."""
    setor_map = {}
    for resp in respostas_qs:
        setor = resp.setor
        if not setor:
            try:
                setor = resp.usuario.perfil.setor
            except Exception:
                setor = None
        if not setor:
            continue
        nivel, cor = _NIVEL_SETOR.get(setor.tipo, ('Outros', '#adb5bd'))
        if setor.pk not in setor_map:
            setor_map[setor.pk] = {
                'label': setor.sigla if setor.sigla else setor.nome,
                'nome': setor.nome,
                'nivel': nivel,
                'cor': cor,
                'valor': Decimal('0'),
            }
        for ri in resp.itens_resposta.all():
            if ri.quantidade and ri.quantidade > 0:
                setor_map[setor.pk]['valor'] += ri.quantidade * ri.item.preco_estimado
    return sorted(setor_map.values(), key=lambda x: x['valor'], reverse=True)


def _get_hom_setor_raiz(setor):
    """Setor usado como chave em HomologacaoSetor para o setor informado."""
    if setor is None:
        return None
    if setor.tipo in ('laboratorio', 'secretaria') and setor.pai_id:
        return setor.pai
    return setor


def _tem_perfil(user, *perfis):
    """Verifica se o usuário tem um dos perfis indicados."""
    if user.is_superuser:
        return True
    try:
        return user.perfil.perfil_tipo in perfis
    except PerfilUsuario.DoesNotExist:
        return False


def gestor_required(view_func):
    """Acesso a: admin, gestor_irp."""
    @functools.wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if _tem_perfil(request.user, 'admin', 'gestor_irp', 'diretor_centro'):
            return view_func(request, *args, **kwargs)
        messages.error(request, 'Acesso restrito a gestores de IRP.')
        return redirect('home')
    return wrapper


def aprovador_setor_required(view_func):
    """Acesso a: admin, aprovador_setor."""
    @functools.wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if _tem_perfil(request.user, 'admin', 'aprovador_setor'):
            return view_func(request, *args, **kwargs)
        messages.error(request, 'Acesso restrito a aprovadores de setor.')
        return redirect('home')
    return wrapper




def licitacao_required(view_func):
    """Acesso a: admin, gestor_irp."""
    @functools.wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if _tem_perfil(request.user, 'admin', 'gestor_irp', 'diretor_centro'):
            return view_func(request, *args, **kwargs)
        messages.error(request, 'Acesso restrito ao acompanhamento de licitação.')
        return redirect('home')
    return wrapper


def gestor_financeiro_required(view_func):
    """Acesso ao módulo de Execução Orçamentária: admin, gestor_financeiro."""
    @functools.wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if _tem_perfil(request.user, 'admin', 'gestor_financeiro'):
            return view_func(request, *args, **kwargs)
        messages.error(request, 'Acesso restrito ao Gestor Financeiro.')
        return redirect('selecionar_modulo')
    return wrapper


def ordenador_required(view_func):
    """Acesso a ações de autorização: admin, ordenador_despesa, gestor_financeiro."""
    @functools.wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if _tem_perfil(request.user, 'admin', 'ordenador_despesa', 'gestor_financeiro'):
            return view_func(request, *args, **kwargs)
        messages.error(request, 'Acesso restrito ao Ordenador de Despesa.')
        return redirect('orcamento_home')
    return wrapper


# ---------------------------------------------------------------------------
# Tela de seleção de módulo
# ---------------------------------------------------------------------------

@login_required
def selecionar_modulo(request):
    """Redireciona para o único módulo disponível ou exibe tela de seleção."""
    try:
        perfil = request.user.perfil
        modulos = perfil.modulos_disponiveis
    except PerfilUsuario.DoesNotExist:
        modulos = ['irp']

    if len(modulos) == 1:
        if modulos[0] == 'orcamento':
            return redirect('orcamento_home')
        return redirect('home')

    return render(request, 'modulos.html', {'modulos': modulos})


# ---------------------------------------------------------------------------
# Views do usuário respondente
# ---------------------------------------------------------------------------

@login_required
def home(request):
    try:
        perfil = request.user.perfil
    except PerfilUsuario.DoesNotExist:
        messages.warning(request, 'Complete seu perfil antes de continuar.')
        return redirect('perfil_editar')

    from django.db.models import Count, Q as Qf
    agora = timezone.now()
    irps_abertas = (
        IRP.objects
        .filter(liberada=True, prazo__gt=agora)
        .annotate(
            itens_ativos=Count('itens', filter=Qf(itens__ativo=True)),
            respostas_enviadas=Count('respostas', filter=Qf(respostas__respondida_em__isnull=False)),
        )
        .order_by('prazo')
    )
    irps_encerradas = IRP.objects.filter(liberada=True, prazo__lte=agora).order_by('-prazo')

    # Apenas respostas efetivamente enviadas (respondida_em preenchido)
    respostas_enviadas = Resposta.objects.filter(
        usuario=request.user, respondida_em__isnull=False
    )
    respostas_dict = {r.irp_id: r for r in respostas_enviadas}

    irps_encerradas_recentes = list(irps_encerradas[:5])

    qt_usuarios = None
    try:
        if request.user.is_superuser or perfil.is_gestor:
            qt_usuarios = User.objects.count()
    except Exception:
        pass

    return render(request, 'core/home.html', {
        'perfil': perfil,
        'qt_abertas':    irps_abertas.count(),
        'qt_encerradas': irps_encerradas.count(),
        'qt_respondidas': respostas_enviadas.count(),
        'qt_total': IRP.objects.count(),
        'qt_usuarios': qt_usuarios,
        'irps_abertas': irps_abertas,
        'irps_encerradas_recentes': irps_encerradas_recentes,
        'respostas_dict': respostas_dict,
        'agora': agora,
    })


@login_required
def irp_list(request):
    try:
        perfil = request.user.perfil
    except PerfilUsuario.DoesNotExist:
        messages.warning(request, 'Complete seu perfil antes de continuar.')
        return redirect('perfil_editar')

    from django.db.models import Count, Q as Qf
    agora = timezone.now()
    irps_abertas = (
        IRP.objects
        .filter(liberada=True, prazo__gt=agora)
        .annotate(
            respostas_enviadas=Count(
                'respostas',
                filter=Qf(respostas__respondida_em__isnull=False),
                distinct=True,
            ),
            itens_ativos=Count(
                'itens',
                filter=Qf(itens__ativo=True),
                distinct=True,
            ),
        )
        .order_by('prazo')
    )
    # Apenas respostas efetivamente enviadas
    respostas_dict = {
        r.irp_id: r
        for r in Resposta.objects.filter(
            usuario=request.user, respondida_em__isnull=False
        )
    }

    context = {
        'irps_abertas': irps_abertas,
        'respostas_dict': respostas_dict,
    }
    return render(request, 'core/irp_list.html', context)


@login_required
def irp_responder(request, pk):
    irp = get_object_or_404(IRP, pk=pk)

    try:
        perfil = request.user.perfil
    except PerfilUsuario.DoesNotExist:
        messages.warning(request, 'Complete seu perfil antes de continuar.')
        return redirect('perfil_editar')

    pode_editar = irp.esta_aberta

    # Bloqueio para usuários sem setor ou setores inapropriados
    if not perfil.setor:
        messages.error(request, 'Você precisa estar vinculado a um setor no seu perfil para responder a uma IRP.')
        return redirect('perfil_editar')

    tipo_setor = perfil.setor.tipo if perfil.setor else ''
    if tipo_setor in ('departamento', 'direcao', 'centro'):
        messages.error(request, 'Departamentos e Direções de Centro não preenchem intenções diretamente. Solicite a um setor administrativamente subordinado.')
        return redirect('irp_list')

    # IRP sem itens: bloquear resposta se ainda aberta
    if pode_editar and not irp.itens.filter(ativo=True).exists():
        messages.error(request, 'Esta IRP ainda não possui itens cadastrados e não pode ser respondida.')
        return redirect('irp_list')

    # Cria ou recupera a resposta do usuário para esta IRP
    resposta, criada = Resposta.objects.get_or_create(
        irp=irp, usuario=request.user,
        defaults={'setor': perfil.setor}
    )
    # Sincroniza o setor caso a resposta existisse antes do perfil ser configurado
    if not criada and resposta.setor is None and perfil.setor:
        resposta.setor = perfil.setor
        resposta.save(update_fields=['setor'])

    itens = list(irp.itens.all())
    respostas_items = {
        ri.item_id: ri
        for ri in RespostaItem.objects.filter(resposta=resposta).select_related('item')
    }

    if request.method == 'POST' and pode_editar:
        try:
            with transaction.atomic():
                for item in itens:
                    qtd_str = request.POST.get(f'qtd_{item.pk}', '').strip().replace(',', '.')
                    obs = request.POST.get(f'obs_{item.pk}', '').strip()
                    try:
                        quantidade = Decimal(qtd_str) if qtd_str else None
                        if quantidade is not None and quantidade <= 0:
                            quantidade = None
                    except InvalidOperation:
                        quantidade = None

                    ri, _ = RespostaItem.objects.get_or_create(resposta=resposta, item=item)
                    ri.quantidade = quantidade
                    ri.observacao = obs
                    ri.save()

                resposta.observacao_geral = request.POST.get('observacao_geral', '').strip()
                from django.utils import timezone as tz
                update_fields = ['observacao_geral', 'atualizada_em']
                if not resposta.respondida_em:
                    resposta.respondida_em = tz.now()
                    update_fields.append('respondida_em')
                resposta.save(update_fields=update_fields)

            messages.success(request, 'Intenções salvas com sucesso!')
            return redirect('irp_list')
        except Exception as e:
            messages.error(request, f'Erro ao salvar: {e}')

    # Monta lista de itens com dados de resposta
    itens_dados = []
    for item in itens:
        ri = respostas_items.get(item.pk)
        itens_dados.append({
            'item': item,
            'quantidade': ri.quantidade if ri else None,
            'observacao': ri.observacao if ri else '',
        })

    ja_respondeu = resposta.itens_resposta.exists()

    context = {
        'irp': irp,
        'resposta': resposta,
        'itens_dados': itens_dados,
        'pode_editar': pode_editar,
        'ja_respondeu': ja_respondeu,
    }
    return render(request, 'core/irp_responder.html', context)


@login_required
def salvar_item_htmx(request, irp_pk, item_pk):
    """Endpoint HTMX: salva a quantidade de um único item."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    irp = get_object_or_404(IRP, pk=irp_pk)
    item = get_object_or_404(Item, pk=item_pk, irp=irp)

    if not irp.esta_aberta:
        return HttpResponse('<span class="text-danger small">IRP encerrada</span>')

    if not perfil.setor:
        return HttpResponse('<span class="text-danger small">Vínculo com setor exigido</span>')

    tipo_setor = perfil.setor.tipo if perfil.setor else ''
    if tipo_setor in ('departamento', 'direcao', 'centro'):
        return HttpResponse('<span class="text-danger small" title="Departamentos não respondem">Bloqueado</span>')

    resposta, _ = Resposta.objects.get_or_create(
        irp=irp, usuario=request.user,
        defaults={'setor': perfil.setor}
    )

    qtd_str = request.POST.get(f'qtd_{item_pk}', '').strip().replace(',', '.')
    obs = request.POST.get(f'obs_{item_pk}', '').strip()

    try:
        quantidade = Decimal(qtd_str) if qtd_str else None
        if quantidade is not None and quantidade <= 0:
            quantidade = None
    except InvalidOperation:
        quantidade = None

    ri, _ = RespostaItem.objects.get_or_create(resposta=resposta, item=item)
    ri.quantidade = quantidade
    ri.observacao = obs
    ri.save()

    # Calcula valor total do item
    if quantidade:
        valor = quantidade * item.preco_estimado
        valor_fmt = _fmt_brl(valor)
        return HttpResponse(
            f'<span class="text-success" title="Salvo">'
            f'<i class="bi bi-check-circle-fill"></i>'
            f'</span>'
            f'<span class="ms-1 text-muted small" id="valor-{item_pk}">{valor_fmt}</span>'
        )
    return HttpResponse(
        f'<span class="text-secondary" title="Salvo (sem intenção)">'
        f'<i class="bi bi-dash-circle"></i>'
        f'</span>'
        f'<span class="ms-1" id="valor-{item_pk}">—</span>'
    )


def _fmt_brl(value):
    if value is None:
        return 'R$ 0,00'
    try:
        v = float(value)
        fmt = f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f'R$ {fmt}'
    except (ValueError, TypeError):
        return 'R$ 0,00'


@login_required
def perfil_editar(request):
    usuario = request.user
    try:
        perfil = usuario.perfil
    except PerfilUsuario.DoesNotExist:
        perfil = None

    todos_setores = Setor.objects.filter(ativo=True).order_by('nome')

    if request.method == 'POST':
        form = UsuarioForm(request.POST, request.FILES, instance=usuario, perfil=perfil)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('irp_list')
    else:
        form = UsuarioForm(instance=usuario, perfil=perfil)

    return render(request, 'core/gestao/usuario_form.html', {
        'form': form,
        'titulo_pagina': 'Meu Perfil',
        'todos_setores': todos_setores,
        'tipo_setor_choices': TIPO_SETOR_CHOICES,
        'perfil': perfil,
        'perfil_por_tipo_json': json.dumps(PERFIL_POR_TIPO_SETOR),
        'modo_perfil': True,
    })


# ---------------------------------------------------------------------------
# Dashboard público
# ---------------------------------------------------------------------------

def dashboard(request):
    irps = IRP.objects.all().order_by('-criada_em')

    # --- GET params (multi-select) ---
    irp_ids_param         = request.GET.getlist('irp')
    pregao_ids_param      = request.GET.getlist('pregao')
    setor_ids_param       = request.GET.getlist('setor')
    solicitante_ids_param = request.GET.getlist('solicitante')
    fase_ids_param        = request.GET.getlist('fase')

    # Resolve pregões → IRP pks
    extra_irp_pks = set()
    for pid in pregao_ids_param:
        try:
            pr = Pregao.objects.select_related('irp').get(pk=pid)
            extra_irp_pks.add(str(pr.irp.pk))
        except Pregao.DoesNotExist:
            pass

    all_irp_pks = set(irp_ids_param) | extra_irp_pks

    FASE_CHOICES = [
        ('em_cadastro',  'Em cadastro'),
        ('respostas',    'Coletando respostas'),
        ('homologacao',  'Em homologação'),
        ('encerrada',    'Encerrada'),
    ]

    # Build irps_para_dados
    irps_para_dados = list(irps)
    if all_irp_pks:
        irps_para_dados = [i for i in irps_para_dados if str(i.pk) in all_irp_pks]
    if fase_ids_param:
        irps_para_dados = [i for i in irps_para_dados if i.fase_atual in fase_ids_param]

    # Setor filter — expand parent setores to include children
    setor_ids_ok = None
    if setor_ids_param:
        setor_ids_ok = set()
        for sid in setor_ids_param:
            try:
                s = Setor.objects.get(pk=sid)
                setor_ids_ok.add(s.pk)
                if s.pai_id is None:
                    setor_ids_ok.update(
                        Setor.objects.filter(pai=s).values_list('pk', flat=True)
                    )
            except Setor.DoesNotExist:
                pass

    # Solicitante filter
    solicitante_pks = None
    if solicitante_ids_param:
        solicitante_pks = set()
        for uid in solicitante_ids_param:
            try:
                solicitante_pks.add(int(uid))
            except (ValueError, TypeError):
                pass

    # Pregões with number assigned (for filter dropdown)
    pregoes = (Pregao.objects.filter(numero__gt='')
               .select_related('irp').order_by('numero'))

    # IRP→Pregao and Pregao→IRP pairs for JS mutual filter
    irp_pregao_pairs = list(
        Pregao.objects.filter(numero__gt='').values('irp_id', 'id')
    )

    # Grouped setores for filter dropdown (only sectors that responded to at least one IRP)
    setores_adm   = Setor.objects.filter(ativo=True, tipo__in=['centro', 'direcao', 'administrativo', 'secretaria'], respostas_setor__respondida_em__isnull=False).distinct().order_by('nome')
    setores_dept  = Setor.objects.filter(ativo=True, tipo__in=['departamento', 'laboratorio'], respostas_setor__respondida_em__isnull=False).distinct().order_by('nome')
    setores_coord = Setor.objects.filter(ativo=True, tipo__in=['coordenacao_g', 'coordenacao_pg'], respostas_setor__respondida_em__isnull=False).distinct().order_by('nome')

    # Solicitantes who responded to any IRP (for filter dropdown)
    solicitantes = (User.objects
                    .filter(respostas__irp__in=irps)
                    .distinct()
                    .select_related('perfil')
                    .order_by('perfil__nome_completo'))

    dados = _calcular_dados(irps_para_dados, setor_ids_ok, solicitante_pks) if irps_para_dados else None

    # Gráfico por setor (mesmo modelo de gestao_resultados, mas agrega todas as IRPs filtradas)
    por_nivel_chart = []
    if irps_para_dados:
        qs_chart = (
            Resposta.objects
            .filter(irp__in=irps_para_dados)
            .select_related('usuario__perfil__setor', 'setor')
            .prefetch_related('itens_resposta__item')
        )
        if solicitante_pks:
            qs_chart = qs_chart.filter(usuario__pk__in=solicitante_pks)
        if setor_ids_ok is not None:
            class _SeitorIter:
                def __init__(self, qs, ids):
                    self._qs, self._ids = qs, ids

                def __iter__(self):
                    for resp in self._qs:
                        s = resp.setor
                        if not s:
                            try:
                                s = resp.usuario.perfil.setor
                            except Exception:
                                s = None
                        if s and s.pk in self._ids:
                            yield resp

            por_nivel_chart = _build_por_nivel_chart(_SeitorIter(qs_chart, setor_ids_ok))
        else:
            por_nivel_chart = _build_por_nivel_chart(qs_chart)

    # ── Gráfico por solicitante ──────────────────────────────────────────────
    por_solicitante_chart = []
    if irps_para_dados:
        sol_map = {}
        qs_sol = (
            Resposta.objects
            .filter(irp__in=irps_para_dados)
            .select_related('usuario__perfil')
            .prefetch_related('itens_resposta__item')
        )
        if setor_ids_ok is not None:
            filtered_resps = []
            for resp in qs_sol:
                s = resp.setor
                if not s:
                    try:
                        s = resp.usuario.perfil.setor
                    except Exception:
                        s = None
                if s and s.pk in setor_ids_ok:
                    filtered_resps.append(resp)
            qs_sol_iter = filtered_resps
        else:
            qs_sol_iter = list(qs_sol)
        if solicitante_pks:
            qs_sol_iter = [r for r in qs_sol_iter if r.usuario_id in solicitante_pks]
        for resp in qs_sol_iter:
            uid = resp.usuario_id
            try:
                nome = resp.usuario.perfil.nome_completo if resp.usuario.perfil.nome_completo else resp.usuario.username
            except Exception:
                nome = resp.usuario.username
            if uid not in sol_map:
                sol_map[uid] = {'nome': nome, 'valor': Decimal('0')}
            for ri in resp.itens_resposta.all():
                if ri.quantidade and ri.quantidade > 0:
                    sol_map[uid]['valor'] += ri.quantidade * ri.item.preco_estimado
        por_solicitante_chart = sorted(sol_map.values(), key=lambda x: x['valor'], reverse=True)

    # ── Tabela por IRP ───────────────────────────────────────────────────────
    por_irp_tabela = []
    if irps_para_dados:
        for irp in irps_para_dados:
            fase = irp.fase_atual
            hom_pendente = fase in ('respostas', 'homologacao')

            # Intencionado: R$, Qtd e Respondentes por item
            item_intenc_v   = {} # item_pk -> R$
            item_intenc_q   = {} # item_pk -> Qtd
            item_resp_users = {} # item_pk -> set(user_pks)
            
            respostas_itens_qs = (RespostaItem.objects
                                  .filter(resposta__irp=irp, quantidade__gt=0, resposta__respondida_em__isnull=False)
                                  .select_related('item', 'resposta'))
            
            for ri in respostas_itens_qs:
                pk = ri.item_id
                val = ri.quantidade * ri.item.preco_estimado
                item_intenc_v[pk] = item_intenc_v.get(pk, Decimal('0')) + val
                item_intenc_q[pk] = item_intenc_q.get(pk, Decimal('0')) + ri.quantidade
                if pk not in item_resp_users:
                    item_resp_users[pk] = set()
                item_resp_users[pk].add(ri.resposta.usuario_id)

            # Homologado: R$ e Qtd por item
            item_hom_v = {}
            item_hom_q = {}
            for hi in (HomologacaoSetorItem.objects
                       .filter(homologacao__irp=irp, homologacao__status='homologada', quantidade_aprovada__isnull=False)
                       .select_related('item')):
                qty = hi.quantidade_aprovada or Decimal('0')
                pk = hi.item_id
                item_hom_v[pk] = item_hom_v.get(pk, Decimal('0')) + qty * hi.item.preco_estimado
                item_hom_q[pk] = item_hom_q.get(pk, Decimal('0')) + qty

            # Licitado: R$ e Qtd por item (baseado na Qtd Homologada e Preço Licitado se houver)
            item_lic_v = {}
            item_lic_q = {}
            p_itens_qs = (PregaoItem.objects
                          .filter(pregao__irp=irp, situacao='licitado')
                          .select_related('item'))
            for pi in p_itens_qs:
                pk = pi.item_id
                q_hom = item_hom_q.get(pk, Decimal('0'))
                if q_hom > 0:
                    p_lic = pi.preco_licitado if pi.preco_licitado else pi.item.preco_estimado
                    item_lic_v[pk] = q_hom * p_lic
                    item_lic_q[pk] = q_hom

            itens_rows = []
            # Listar todos os itens da IRP que tiveram intenção
            todos_itens_irp = irp.itens.all().order_by('numero')
            for item_obj in todos_itens_irp:
                pk = item_obj.pk
                r_intenc = item_intenc_v.get(pk, Decimal('0'))
                if r_intenc == 0 and pk not in item_hom_v:
                    continue # Pula itens sem qualquer movimento
                
                itens_rows.append({
                    'item': item_obj,
                    'n_respondentes': len(item_resp_users.get(pk, [])),
                    'r_intenc': r_intenc,
                    'q_intenc': item_intenc_q.get(pk, Decimal('0')),
                    'r_hom': item_hom_v.get(pk, Decimal('0')),
                    'q_hom': item_hom_q.get(pk, Decimal('0')),
                    'r_lic': item_lic_v.get(pk, Decimal('0')),
                    'q_lic': item_lic_q.get(pk, Decimal('0')),
                })

            if itens_rows:
                por_irp_tabela.append({
                    'irp': irp,
                    'n_respostas': Resposta.objects.filter(irp=irp, respondida_em__isnull=False).values('usuario').distinct().count(),
                    'r_intenc': sum(r['r_intenc'] for r in itens_rows),
                    'q_intenc': sum(r['q_intenc'] for r in itens_rows),
                    'r_hom': sum(r['r_hom'] for r in itens_rows),
                    'q_hom': sum(r['q_hom'] for r in itens_rows),
                    'r_lic': sum(r['r_lic'] for r in itens_rows),
                    'q_lic': sum(r['q_lic'] for r in itens_rows),
                    'hom_pendente': hom_pendente,
                    'itens': itens_rows,
                })
        por_irp_tabela.sort(key=lambda x: x['r_intenc'], reverse=True)

    return render(request, 'core/dashboard.html', {
        'irps': irps,
        'pregoes': pregoes,
        'irp_pregao_pairs': irp_pregao_pairs,
        'setores_adm': setores_adm,
        'setores_dept': setores_dept,
        'setores_coord': setores_coord,
        'solicitantes': solicitantes,
        'irp_ids_param': irp_ids_param,
        'pregao_ids_param': pregao_ids_param,
        'setor_ids_param': setor_ids_param,
        'solicitante_ids_param': solicitante_ids_param,
        'fase_ids_param': fase_ids_param,
        'fase_choices': FASE_CHOICES,
        'dados': dados,
        'por_nivel_chart': por_nivel_chart,
        'por_solicitante_chart': por_solicitante_chart,
        'por_irp_tabela': por_irp_tabela,
    })


def _calcular_dados(irps, setor_ids_ok=None, solicitante_pks=None):
    # Busca todas as respostas com fallback de setor via perfil
    respostas_qs = (
        Resposta.objects
        .filter(irp__in=irps, respondida_em__isnull=False)
        .select_related('usuario__perfil__setor', 'setor')
        .prefetch_related('itens_resposta__item')
    )
    if solicitante_pks:
        respostas_qs = respostas_qs.filter(usuario__pk__in=solicitante_pks)

    # Agrega por setor resolvido (com fallback de perfil)
    setor_data = {}   # setor_pk → {setor, respondentes, r_intenc, itens_set, hom_raiz}
    total_r_intenc = Decimal('0')
    total_respostas = 0      # Resposta objects (uma por IRP por usuário)
    usuarios_set = set()     # usuários únicos (respondentes)

    for resp in respostas_qs:
        setor = resp.setor
        if not setor:
            try:
                setor = resp.usuario.perfil.setor
            except Exception:
                setor = None

        if setor_ids_ok is not None:
            if setor is None or setor.pk not in setor_ids_ok:
                continue

        total_respostas += 1
        usuarios_set.add(resp.usuario_id)
        r_resp = Decimal('0')
        itens_resp = set()
        for ri in resp.itens_resposta.all():
            if ri.quantidade and ri.quantidade > 0:
                r_resp += ri.quantidade * ri.item.preco_estimado
                itens_resp.add(ri.item_id)

        total_r_intenc += r_resp

        if setor:
            hom_raiz = _get_hom_setor_raiz(setor)
            if setor.pk not in setor_data:
                setor_data[setor.pk] = {
                    'setor': setor,
                    'respondentes': 0,
                    'r_intenc': Decimal('0'),
                    'itens_set': set(),
                    'hom_raiz': hom_raiz,
                }
            setor_data[setor.pk]['respondentes'] += 1
            setor_data[setor.pk]['r_intenc'] += r_resp
            setor_data[setor.pk]['itens_set'].update(itens_resp)

    # Hom por hom_raiz único
    hom_raiz_pks = {sp['hom_raiz'].pk for sp in setor_data.values() if sp['hom_raiz']}
    raiz_hom_data = {}  # raiz_pk → {r_hom}
    for raiz_pk in hom_raiz_pks:
        r_hom = Decimal('0')
        for hom in HomologacaoSetor.objects.filter(irp__in=irps, setor_raiz_id=raiz_pk, status='homologada'):
            for hi in (HomologacaoSetorItem.objects
                       .filter(homologacao=hom, quantidade_aprovada__isnull=False)
                       .select_related('item')):
                r_hom += (hi.quantidade_aprovada or 0) * hi.item.preco_estimado
        raiz_hom_data[raiz_pk] = {'r_hom': r_hom}

    # total_r_hom/total_r_lic: somas diretas dos itens homologados/licitados das IRPs
    total_r_hom = Decimal('0')
    total_r_lic = Decimal('0')
    
    # Cache licitado pks e preços
    lic_map = {} # item_pk -> preco_licitado or preco_estimado
    for pi in PregaoItem.objects.filter(pregao__irp__in=irps, situacao='licitado').select_related('item'):
        lic_map[pi.item_id] = pi.preco_licitado if pi.preco_licitado else pi.item.preco_estimado

    for hi in (HomologacaoSetorItem.objects
               .filter(homologacao__irp__in=irps, homologacao__status='homologada',
                       quantidade_aprovada__isnull=False)
               .select_related('item')):
        qty = hi.quantidade_aprovada or Decimal('0')
        total_r_hom += qty * hi.item.preco_estimado
        if hi.item_id in lic_map:
            total_r_lic += qty * lic_map[hi.item_id]

    # Agrupa por nível (Adm / Dept+Labs / Coord)
    NIVEL_GRUPOS = [
        ('Setores Administrativos', {'centro', 'direcao', 'administrativo', 'secretaria'}),
        ('Departamentos e Laboratórios', {'departamento', 'laboratorio'}),
        ('Coordenações', {'coordenacao_g', 'coordenacao_pg'}),
    ]

    por_setor = []

    for grupo_nome, tipos in NIVEL_GRUPOS:
        g_respondentes = 0
        g_itens = 0
        g_r_intenc = Decimal('0')
        g_r_hom = Decimal('0')
        por_sub = []
        seen_raiz_pks = set()

        for sp in sorted(setor_data.values(), key=lambda x: x['r_intenc'], reverse=True):
            if sp['setor'].tipo not in tipos:
                continue
            g_respondentes += sp['respondentes']
            g_itens += len(sp['itens_set'])
            g_r_intenc += sp['r_intenc']
            hr = sp['hom_raiz']
            if hr and hr.pk not in seen_raiz_pks:
                seen_raiz_pks.add(hr.pk)
                hd = raiz_hom_data.get(hr.pk, {})
                g_r_hom += hd.get('r_hom', Decimal('0'))
            por_sub.append({
                'setor': sp['setor'],
                'respondentes': sp['respondentes'],
                'itens': len(sp['itens_set']),
                'r_intenc': sp['r_intenc'],
            })

        if por_sub:
            por_setor.append({
                'grupo_nome': grupo_nome,
                'respondentes': g_respondentes,
                'itens': g_itens,
                'r_intenc': g_r_intenc,
                'r_hom': g_r_hom,
                'por_subsetor': por_sub,
            })

    por_setor.sort(key=lambda x: x['r_intenc'], reverse=True)

    return {
        'total_respondentes': len(usuarios_set),
        'total_respostas': total_respostas,
        'total_itens_irp': sum(irp.itens.count() for irp in irps),
        'total_r_intenc': total_r_intenc,
        'total_r_hom': total_r_hom,
        'total_r_lic': total_r_lic,
        'por_setor': por_setor,
    }


# ---------------------------------------------------------------------------
# Homologação por Setor Raiz
# ---------------------------------------------------------------------------

@aprovador_setor_required
def homologacao_list(request):
    """Lista IRPs com respostas do grupo do usuário pendentes de homologação."""
    from django.db.models import Count, Q as Qf
    perfil = request.user.perfil
    setor_raiz = perfil.get_setor_raiz()
    is_supervisor = perfil.perfil_tipo == 'admin'

    if not setor_raiz and not is_supervisor:
        return render(request, 'core/homologacao_list.html', {
            'irps_com_status': [],
            'setor_raiz': None,
            'is_admin_view': False,
        })

    # Definir quais setores filtrar
    if is_supervisor:
        # Admin vê tudo o que tem resposta
        todos_setores = None
    elif setor_raiz and setor_raiz.tipo == 'direcao':
        # DC homologa os setores de bypass (administrativo e coordenações)
        todos_setores = list(Setor.objects.filter(tipo__in=['administrativo', 'coordenacao_g', 'coordenacao_pg'], ativo=True))
    else:
        # Usuário comum vê seu grupo (raiz + filhos)
        subsetores = Setor.objects.filter(pai=setor_raiz, ativo=True)
        todos_setores = [setor_raiz] + list(subsetores)

    # Buscar IRPs
    filter_kwargs = {
        'itens__respostas_item__quantidade__gt': 0,
        'itens__respostas_item__resposta__respondida_em__isnull': False,
    }
    if todos_setores:
        filter_kwargs['itens__respostas_item__resposta__setor__in'] = todos_setores

    irps_com_intencao = (IRP.objects
                         .filter(**filter_kwargs)
                         .distinct()
                         .annotate(
                             itens_ativos=Count('itens', filter=Qf(itens__ativo=True), distinct=True),
                             total_respostas=Count('respostas', filter=Qf(respostas__respondida_em__isnull=False), distinct=True),
                         )
                         .order_by('-criada_em'))

    irps_com_status = []
    for irp in irps_com_intencao:
        # Se supervisor, o status de "homologação" é um resumo (homologada se todos os envolvidos homologaram)
        if is_supervisor:
            # Status real: verifica as HomologacaoSetor existentes
            homs = HomologacaoSetor.objects.filter(irp=irp)
            total_homs = homs.count()
            if total_homs == 0:
                hstatus = 'pendente'
            elif homs.filter(status='homologada').count() == total_homs:
                hstatus = 'homologada'
            elif homs.filter(status='rejeitada').exists():
                hstatus = 'rejeitada'
            else:
                hstatus = 'pendente'

            hom_fake = type('FakeHom', (), {'status': hstatus})()
        else:
            hom_fake = HomologacaoSetor.objects.filter(irp=irp, setor_raiz=setor_raiz).first()

        has_pregao = Pregao.objects.filter(irp=irp).exists()
        irps_com_status.append({
            'irp': irp,
            'homologacao': hom_fake,
            'has_pregao': has_pregao,
            # bloqueia reabrir se já encaminhada para licitação (pregão criado)
            'em_licitacao': has_pregao,
        })

    return render(request, 'core/homologacao_list.html', {
        'irps_com_status': irps_com_status,
        'setor_raiz': setor_raiz,
        'is_admin_view': is_supervisor,
    })


@aprovador_setor_required
def homologar_setor(request, irp_pk):
    """Tela de homologação: mostra respostas do setor raiz e subsetores."""
    irp = get_object_or_404(IRP, pk=irp_pk)
    perfil = request.user.perfil
    is_supervisor = perfil.perfil_tipo == 'admin'

    # Se supervisor passar setor_id via GET, usamos esse setor.
    # Caso contrário, tenta descobrir o setor_raiz do próprio perfil.
    setor_id_manual = request.GET.get('setor_id')
    if is_supervisor and setor_id_manual:
        setor_raiz = get_object_or_404(Setor, pk=setor_id_manual)
    else:
        setor_raiz = perfil.get_setor_raiz()

    if not setor_raiz:
        # Se for supervisor, tenta pegar o setor 'direcao' como padrão
        if is_supervisor:
            setor_raiz = Setor.objects.filter(tipo='direcao').first()
        
        if not setor_raiz:
            if is_supervisor:
                messages.info(request, "Direcionado para Resultados (Supervisão), selecione o setor nos filtros se desejar homologação específica.")
                return redirect('gestao_resultados', irp_pk=irp_pk)

            messages.error(request, 'Você não possui setor de lotação definido.')
            return redirect('homologacao_list')

    # Todos os setores do grupo (raiz + filhos)
    # DC homologa os setores de bypass diretamente
    if setor_raiz.tipo == 'direcao':
        todos_setores = list(Setor.objects.filter(tipo__in=['administrativo', 'coordenacao_g', 'coordenacao_pg'], ativo=True))
    else:
        subsetores = list(Setor.objects.filter(pai=setor_raiz, ativo=True))
        todos_setores = [setor_raiz] + subsetores

    # Respostas do grupo para esta IRP
    respostas = (Resposta.objects
                 .filter(irp=irp, setor__in=todos_setores)
                 .select_related('usuario__perfil', 'setor')
                 .order_by('setor__nome'))

    # Setores que efetivamente possuem respostas (para o filtro de subsetores)
    _setores_com_resp_pks = set(respostas.values_list('setor_id', flat=True))
    setores_com_respostas = [s for s in todos_setores if s.pk in _setores_com_resp_pks]

    # Itens intencionados por ao menos uma resposta enviada do grupo
    itens_com_intencao_ids = (
        RespostaItem.objects
        .filter(
            item__irp=irp,
            quantidade__gt=0,
            resposta__respondida_em__isnull=False,
            resposta__setor__in=todos_setores,
        )
        .values_list('item_id', flat=True)
        .distinct()
    )
    itens = irp.itens.filter(ativo=True, pk__in=itens_com_intencao_ids).order_by('numero')

    # Monta dict: {resposta_id: {item_id: RespostaItem}}
    from collections import defaultdict
    ri_map = defaultdict(dict)
    for ri in RespostaItem.objects.filter(resposta__in=respostas).select_related('item'):
        ri_map[ri.resposta_id][ri.item_id] = ri

    homologacao, _ = HomologacaoSetor.objects.get_or_create(
        irp=irp, setor_raiz=setor_raiz
    )

    # Mapa de ajustes já salvos: {item_id: HomologacaoSetorItem}
    hom_itens_map = {
        hi.item_id: hi
        for hi in HomologacaoSetorItem.objects.filter(homologacao=homologacao)
    }

    if request.method == 'POST' and homologacao.status == 'pendente':
        status_novo = request.POST.get('status')
        observacao = request.POST.get('observacao', '').strip()
        if status_novo in ('homologada', 'rejeitada'):
            # Salva ajustes item a item
            for item in itens:
                qty_str = request.POST.get(f'qty_{item.pk}', '').strip()
                obs_str = request.POST.get(f'obs_{item.pk}', '').strip()
                hom_item = hom_itens_map.get(item.pk)
                if not hom_item:
                    hom_item = HomologacaoSetorItem(homologacao=homologacao, item=item)
                try:
                    from decimal import Decimal, InvalidOperation
                    hom_item.quantidade_aprovada = Decimal(qty_str) if qty_str else None
                except InvalidOperation:
                    hom_item.quantidade_aprovada = None
                hom_item.observacao = obs_str
                hom_item.save()

            homologacao.status = status_novo
            homologacao.observacao = observacao
            homologacao.homologado_por = request.user
            homologacao.homologado_em = timezone.now()
            homologacao.save()
            label = 'homologada' if status_novo == 'homologada' else 'rejeitada'
            messages.success(request, f'IRP {label} para o setor {setor_raiz.nome}.')
            return redirect('homologacao_list')

    pode_reabrir_hom = (
        homologacao.status != 'pendente'
        and not Pregao.objects.filter(irp=irp).exists()
    )

    # Modo somente leitura: explicitamente via ?readonly=1 OU homologação já decidida
    readonly = request.GET.get('readonly') == '1' or homologacao.status != 'pendente'

    return render(request, 'core/homologacao.html', {
        'irp': irp,
        'setor_raiz': setor_raiz,
        'todos_setores': todos_setores,
        'setores_com_respostas': setores_com_respostas,
        'respostas': respostas,
        'itens': itens,
        'ri_map': ri_map,
        'homologacao': homologacao,
        'hom_itens_map': hom_itens_map,
        'pode_reabrir_hom': pode_reabrir_hom,
        'readonly': readonly,
        'agora': timezone.now(),
    })


@aprovador_setor_required
def reabrir_homologacao(request, irp_pk):
    """Reabre uma homologação já decidida, desde que no prazo e o CT não tenha atuado."""
    if request.method != 'POST':
        return HttpResponse(status=405)
    irp = get_object_or_404(IRP, pk=irp_pk)

    if Pregao.objects.filter(irp=irp).exists():
        messages.error(request, 'IRP já foi encaminhada para licitação. Não é possível reabrir a homologação.')
        return redirect('homologacao_list')

    perfil = request.user.perfil
    is_supervisor = perfil.perfil_tipo == 'admin'

    if is_supervisor:
        # Admin reabre todas as homologações da IRP de uma vez
        count = HomologacaoSetor.objects.filter(irp=irp).update(
            status='pendente',
            observacao='',
            homologado_por=None,
            homologado_em=None,
        )
        messages.success(request, f'Homologação(ões) reaberta(s) para "{irp.titulo}" ({count} setor(es)).')
        return redirect('homologacao_list')

    setor_raiz = perfil.get_setor_raiz()
    if not setor_raiz:
        messages.error(request, 'Setor de lotação não definido.')
        return redirect('homologacao_list')

    homologacao = get_object_or_404(HomologacaoSetor, irp=irp, setor_raiz=setor_raiz)
    homologacao.status = 'pendente'
    homologacao.observacao = ''
    homologacao.homologado_por = None
    homologacao.homologado_em = None
    homologacao.save()
    messages.success(request, f'Homologação reaberta para {setor_raiz.nome}.')
    return redirect('homologacao_list')


@aprovador_setor_required
def encaminhar_licitacao(request, irp_pk):
    """Encaminha uma IRP para a fila de licitação (cria o Pregão)."""
    if request.method != 'POST':
        return HttpResponse(status=405)
    irp = get_object_or_404(IRP, pk=irp_pk)

    if irp.fase_atual != 'encerrada':
        messages.error(request, 'A IRP ainda não concluiu a fase de homologação e não pode ser encaminhada.')
        return redirect('homologacao_list')

    pregao, created = Pregao.objects.get_or_create(irp=irp)
    if created:
        messages.success(request, f'IRP "{irp.titulo}" encaminhada para licitação.')
    else:
        messages.info(request, f'IRP "{irp.titulo}" já havia sido encaminhada para licitação.')
    return redirect('homologacao_list')


# ---------------------------------------------------------------------------
# Aprovação pelo CT
# ---------------------------------------------------------------------------



# ---------------------------------------------------------------------------
# Licitação / Pregão
# ---------------------------------------------------------------------------

@licitacao_required
def licitacao_list(request):
    """Lista IRPs encaminhadas explicitamente para licitação."""
    rows = [
        {'irp': pregao.irp, 'pregao': pregao}
        for pregao in Pregao.objects.select_related('irp').order_by('-irp__criada_em')
    ]
    return render(request, 'core/licitacao_list.html', {'rows': rows})


@licitacao_required
def licitacao_detalhe(request, irp_pk):
    """Detalhe dos itens do pregão para uma IRP aprovada."""
    irp = get_object_or_404(IRP, pk=irp_pk)
    if irp.fase_atual != 'encerrada':
        messages.error(request, 'Esta IRP ainda não concluiu a fase de homologação.')
        return redirect('licitacao_list')
    pregao, _ = Pregao.objects.get_or_create(irp=irp)
    # Apenas itens intencionados por ao menos uma resposta enviada
    itens_com_intencao_ids = (
        RespostaItem.objects
        .filter(
            item__irp=irp,
            quantidade__gt=0,
            resposta__respondida_em__isnull=False,
        )
        .values_list('item_id', flat=True)
        .distinct()
    )
    itens = irp.itens.filter(ativo=True, pk__in=itens_com_intencao_ids).order_by('numero')
    # Mapa item_id → quantidade homologada (soma de todas as homologações aprovadas da IRP)
    from django.db.models import Sum
    qty_hom_map = {
        row['item_id']: row['total']
        for row in HomologacaoSetorItem.objects
            .filter(homologacao__irp=irp, homologacao__status='homologada',
                    quantidade_aprovada__isnull=False)
            .values('item_id')
            .annotate(total=Sum('quantidade_aprovada'))
    }

    itens_data = []
    for item in itens:
        pi, _ = PregaoItem.objects.get_or_create(pregao=pregao, item=item)
        # Quantidade homologada; fallback para intencionada se não houver homologação
        qty_hom = qty_hom_map.get(item.pk)
        if qty_hom is None:
            qty_hom = sum(
                ri.quantidade or 0
                for ri in item.respostas_item.filter(quantidade__gt=0)
            )
        # Preço efetivo: licitado se disponível, senão estimado
        preco_ef = pi.preco_licitado if pi.preco_licitado else item.preco_estimado
        itens_data.append({
            'item': item,
            'pi': pi,
            'total_qty': qty_hom,
            'valor_total': qty_hom * preco_ef if qty_hom else None,
        })
    return render(request, 'core/licitacao_detalhe.html', {
        'irp': irp,
        'pregao': pregao,
        'itens_data': itens_data,
    })


@licitacao_required
def licitacao_salvar_pregao(request, pregao_pk):
    """AJAX: salva campo do pregão (numero, link, datas, status)."""
    if request.method != 'POST':
        return HttpResponse(status=405)
    pregao = get_object_or_404(Pregao, pk=pregao_pk)
    campo = request.POST.get('campo', '')
    valor = request.POST.get('valor', '').strip()
    if campo == 'numero':
        pregao.numero = valor
    elif campo == 'link':
        pregao.link_acompanhamento = valor
    elif campo == 'data_publicacao':
        try:
            pregao.data_publicacao = valor if valor else None
        except Exception:
            return HttpResponse('Data inválida', status=400)
    elif campo == 'data_homologacao':
        try:
            pregao.data_homologacao = valor if valor else None
        except Exception:
            return HttpResponse('Data inválida', status=400)
    elif campo == 'status' and valor in [s[0] for s in Pregao.STATUS_CHOICES]:
        pregao.status = valor
    elif campo == 'validar_status':
        # Valor deve ser o novo status a ser validado
        if valor not in [s[0] for s in Pregao.STATUS_CHOICES]:
            return HttpResponse('Status inválido', status=400)
        
        itens = pregao.itens_pregao.all()
        if valor == 'em_preparacao':
            itens.update(situacao='em_preparacao')
        elif valor == 'em_licitacao':
            # Apenas itens que NÃO são licitado/não_licitado vão para em_licitacao
            itens.exclude(situacao__in=['licitado', 'nao_licitado']).update(situacao='em_licitacao')
        elif valor == 'homologado':
            # BLOQUEIO: Não homologar se houver itens em preparação ou em licitação
            itens_pendentes = itens.exclude(situacao__in=['licitado', 'nao_licitado'])
            if itens_pendentes.exists():
                nums = list(itens_pendentes.values_list('item__numero', flat=True))
                import json
                return HttpResponse(
                    json.dumps({'error': 'itens_pendentes', 'itens': nums}),
                    content_type='application/json', status=400
                )
        # Se passou na validação ou não é homologado, salva o status
        pregao.status = valor
        pregao.save()
        return HttpResponse(status=204)
        
    else:
        return HttpResponse(status=400)

    # Para os campos individuais (numero, link, etc) que não caíram no validar_status
    pregao.save()
    return HttpResponse(status=204)


@licitacao_required
def licitacao_salvar_item(request, item_pk):
    """AJAX: salva campo do PregaoItem (preco_licitado, situacao)."""
    if request.method != 'POST':
        return HttpResponse(status=405)
    pi = get_object_or_404(PregaoItem, item_id=item_pk)
    campo = request.POST.get('campo', '')
    valor = request.POST.get('valor', '').strip()
    if campo == 'preco_licitado':
        try:
            pi.preco_licitado = Decimal(valor.replace(',', '.')) if valor else None
        except InvalidOperation:
            return HttpResponse('Valor inválido', status=400)
    elif campo == 'situacao' and valor in [s[0] for s in PregaoItem.SITUACAO_CHOICES]:
        pi.situacao = valor
    else:
        return HttpResponse(status=400)
    pi.save()
    return HttpResponse(status=204)


@licitacao_required
def licitacao_exportar_itens(request, pregao_pk):
    """Gera e devolve a planilha de acompanhamento de itens da licitação em duas abas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.db.models import Sum, Q

    pregao = get_object_or_404(Pregao, pk=pregao_pk)
    if pregao.status != 'homologado':
        from django.contrib import messages
        messages.error(request, 'A planilha só pode ser exportada após a homologação.')
        return redirect('licitacao_list')
        
    irp = pregao.irp
    # Agregação por item para a síntese
    items_agg = irp.itens.annotate(
        q_intenc=Sum('respostas_item__quantidade', filter=Q(respostas_item__resposta__respondida_em__isnull=False)),
        q_hom=Sum('homologacoes_setor__quantidade_aprovada', filter=Q(homologacoes_setor__homologacao__status='homologada'))
    ).order_by('numero')

    # Mapeamento do PregaoItem
    pi_map = {pi.item_id: pi for pi in pregao.itens_pregao.all()}

    wb = openpyxl.Workbook()
    
    # Estilos
    header_font    = Font(bold=True, color='FFFFFF', size=10)
    header_fill    = PatternFill('solid', fgColor='1D4ED8')
    center         = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left           = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right          = Alignment(horizontal='right',  vertical='center', wrap_text=True)
    thin           = Side(style='thin', color='CCCCCC')
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Aba 1: Síntese da Licitação ──
    ws1 = wb.active
    ws1.title = "Síntese da Licitação"
    
    headers1 = ['ITEM', 'UND', 'RUBRICA', 'N° DFD', 'CÓDIGO CATMAT', 'DISCRIMINAÇÃO DE MATERIAL', 'VALOR', 'QTE INTENCIONADA', 'QTE HOMOLOGADA', 'Preço Licitado', 'Valor Licitado', 'Situação']
    ws1.append(headers1)
    for col, _ in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws1.row_dimensions[1].height = 30

    row_num1 = 2
    for item in items_agg:
        rubrica = item.get_rubrica_display() if item.rubrica else ''
        pi = pi_map.get(item.pk)
        
        p_lic = pi.preco_licitado if pi and pi.preco_licitado else None
        qty_hom = item.q_hom or Decimal('0')
        v_lic = qty_hom * p_lic if p_lic else None
        
        linha1 = [
            item.numero,
            item.unidade,
            rubrica,
            item.numero_dfd,
            item.codigo_catmat,
            item.descricao,
            float(item.preco_estimado or 0),
            float(item.q_intenc or 0),
            float(qty_hom),
            float(p_lic) if p_lic else '',
            float(v_lic) if v_lic else '',
            pi.get_situacao_display() if pi else '—',
        ]
        ws1.append(linha1)
        for col, val in enumerate(linha1, 1):
            cell = ws1.cell(row=row_num1, column=col)
            cell.border = border
            cell.alignment = center if col in (1, 2, 4, 5, 8, 9, 12) else (right if col in (7, 10, 11) else left)
            if col in (7, 8, 9, 10, 11):
                cell.number_format = '#,##0.00'
        row_num1 += 1

    larguras1 = [8, 10, 25, 12, 12, 50, 14, 16, 16, 16, 18, 15]
    for i, w in enumerate(larguras1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'

    # ── Aba 2: Respostas do CT ──
    ws2 = wb.create_sheet(title="Respostas do CT")
    headers2 = [
        'IRP', 'Setor', 'Nome do Respondente', 'Matrícula',
        'Nº Item', 'Rubrica', 'Descrição do Item', 'Unidade',
        'Preço Estimado (R$)', 'Quantidade Intencionada',
        'Valor Total (R$)', 'Observação', 'Última Atualização',
    ]
    ws2.append(headers2)
    for col, _ in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[1].height = 30

    respostas = (
        Resposta.objects
        .filter(irp=irp, respondida_em__isnull=False)
        .select_related('usuario__perfil', 'setor')
        .prefetch_related('itens_resposta__item')
        .order_by('setor__nome')
    )

    row_num2 = 2
    for resp in respostas:
        try:
            p = resp.usuario.perfil
            nome, matricula = p.nome_completo, p.matricula or ''
        except Exception:
            nome, matricula = resp.usuario.get_full_name() or resp.usuario.username, ''

        for ri in resp.itens_resposta.filter(quantidade__gt=0).order_by('item__numero'):
            valor = ri.quantidade * ri.item.preco_estimado
            linha2 = [
                irp.titulo,
                resp.setor.nome if resp.setor else '',
                nome, matricula,
                ri.item.numero,
                ri.item.get_rubrica_display() if ri.item.rubrica else '',
                ri.item.descricao,
                ri.item.unidade,
                float(ri.item.preco_estimado or 0),
                float(ri.quantidade or 0),
                float(valor or 0),
                ri.observacao or '',
                resp.atualizada_em.strftime('%d/%m/%Y %H:%M'),
            ]
            ws2.append(linha2)
            for col, _ in enumerate(linha2, 1):
                cell = ws2.cell(row=row_num2, column=col)
                cell.border = border
                cell.alignment = center if col in (1, 5, 8, 9, 10, 11) else left
                if col in (9, 10, 11):
                    cell.number_format = '#,##0.00'
            row_num2 += 1

    larguras2 = [25, 25, 30, 15, 8, 20, 50, 10, 15, 16, 16, 30, 18]
    for i, w in enumerate(larguras2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"licitacao_{pregao.numero or pregao.pk}.xlsx".replace('/', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# Gestão — IRP
# ---------------------------------------------------------------------------

@gestor_required
def gestao_home(request):
    agora = timezone.now()
    return render(request, 'core/gestao/home.html', {
        'total_irps': IRP.objects.count(),
        'irps_abertas': IRP.objects.filter(liberada=True, prazo__gt=agora).count(),
        'total_usuarios': User.objects.count(),
        'total_setores': Setor.objects.filter(ativo=True).count(),
        'total_respostas': Resposta.objects.count(),
    })


@gestor_required
def gestao_irp_list(request):
    from django.db.models import Count, Q as Qf
    irps = (IRP.objects.all()
            .annotate(
                itens_ativos=Count('itens', filter=Qf(itens__ativo=True), distinct=True),
                total_respostas=Count('respostas', filter=Qf(respostas__respondida_em__isnull=False), distinct=True),
            )
            .order_by('-criada_em'))
    return render(request, 'core/gestao/irp_list.html', {'irps': irps})


@gestor_required
def gestao_irp_create(request):
    if request.method == 'POST':
        form = IRPForm(request.POST)
        if form.is_valid():
            irp = form.save(commit=False)
            irp.criada_por = request.user
            irp.save()
            messages.success(request, f'IRP "{irp.titulo}" criada. Adicione os itens abaixo.')
            return redirect('gestao_item_list', irp_pk=irp.pk)
    else:
        form = IRPForm()
    return render(request, 'core/gestao/irp_form.html', {
        'form': form, 'titulo_pagina': 'Nova IRP'
    })


@gestor_required
def gestao_irp_edit(request, pk):
    irp = get_object_or_404(IRP, pk=pk)
    if irp.liberada:
        messages.error(request, f'A IRP "{irp.titulo}" está em fluxo e não pode ser editada. Interrompa-a primeiro.')
        return redirect('gestao_irp_list')
    if request.method == 'POST':
        form = IRPForm(request.POST, instance=irp)
        if form.is_valid():
            form.save()
            messages.success(request, 'IRP atualizada.')
            return redirect('gestao_irp_list')
    else:
        form = IRPForm(instance=irp)
    return render(request, 'core/gestao/irp_form.html', {
        'form': form, 'irp': irp, 'titulo_pagina': 'Editar IRP'
    })


@gestor_required
def gestao_irp_liberar(request, pk):
    irp = get_object_or_404(IRP, pk=pk)
    if request.method == 'POST':
        if not irp.itens.filter(ativo=True).exists():
            messages.error(request, f'A IRP "{irp.titulo}" não possui itens cadastrados. Adicione itens antes de liberar.')
        else:
            irp.liberada = True
            irp.save(update_fields=['liberada'])
            messages.success(request, f'IRP "{irp.titulo}" liberada para receber respostas.')
    return redirect('gestao_irp_list')


@gestor_required
def gestao_irp_interromper(request, pk):
    irp = get_object_or_404(IRP, pk=pk)
    if request.method == 'POST':
        if irp.fase_atual == 'encerrada':
            messages.error(request, f'A IRP "{irp.titulo}" já está encerrada e não pode ser interrompida.')
        elif not irp.liberada:
            messages.error(request, f'A IRP "{irp.titulo}" não está liberada.')
        else:
            irp.liberada = False
            irp.save(update_fields=['liberada'])
            messages.warning(request, f'IRP "{irp.titulo}" interrompida. Faça os ajustes necessários e libere novamente.')
    return redirect('gestao_irp_list')




@gestor_required
def gestao_irp_delete(request, pk):
    irp = get_object_or_404(IRP, pk=pk)
    if request.method == 'POST':
        titulo = irp.titulo
        irp.delete()
        messages.success(request, f'IRP "{titulo}" excluída.')
    return redirect('gestao_irp_list')


# ---------------------------------------------------------------------------
# Gestão — Itens
# ---------------------------------------------------------------------------

@gestor_required
def gestao_item_list(request, irp_pk):
    irp = get_object_or_404(IRP, pk=irp_pk)
    if irp.liberada:
        messages.error(request, f'A IRP "{irp.titulo}" está em fluxo. Interrompa-a para gerenciar os itens.')
        return redirect('gestao_irp_list')
    itens = irp.itens.all()
    return render(request, 'core/gestao/item_list.html', {
        'irp': irp, 'itens': itens,
    })


@gestor_required
def gestao_item_create(request, irp_pk):
    irp = get_object_or_404(IRP, pk=irp_pk)
    if irp.fase_atual == 'encerrada':
        messages.error(request, f'A IRP "{irp.titulo}" está encerrada. Não é possível adicionar itens.')
        return redirect('gestao_item_list', irp_pk=irp.pk)
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.irp = irp
            item.save()
            messages.success(request, f'Item {item.numero} adicionado.')
            return redirect('gestao_item_list', irp_pk=irp.pk)
    else:
        form = ItemForm()
    return render(request, 'core/gestao/item_form.html', {
        'form': form, 'irp': irp, 'item': None,
    })


@gestor_required
def gestao_item_edit(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if item.irp.fase_atual == 'encerrada':
        messages.error(request, f'A IRP "{item.irp.titulo}" está encerrada e não pode ser editada.')
        return redirect('gestao_item_list', irp_pk=item.irp.pk)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Item atualizado.')
            return redirect('gestao_item_list', irp_pk=item.irp.pk)
    else:
        form = ItemForm(instance=item)
    return render(request, 'core/gestao/item_form.html', {'form': form, 'item': item})


@gestor_required
@gestor_required
def gestao_item_toggle(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if item.irp.fase_atual == 'encerrada':
        messages.error(request, f'A IRP "{item.irp.titulo}" está encerrada.')
        return redirect('gestao_item_list', irp_pk=item.irp.pk)
    if request.method == 'POST':
        item.ativo = not item.ativo
        item.save()
        estado = 'ativado' if item.ativo else 'desativado'
        messages.success(request, f'Item {item.numero} {estado}.')
    return redirect('gestao_item_list', irp_pk=item.irp.pk)


@gestor_required
def gestao_item_delete(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if item.irp.fase_atual == 'encerrada':
        messages.error(request, f'A IRP "{item.irp.titulo}" está encerrada.')
        return redirect('gestao_item_list', irp_pk=item.irp.pk)
    irp_pk = item.irp.pk
    if request.method == 'POST':
        item.delete()
        messages.success(request, 'Item excluído.')
    return redirect('gestao_item_list', irp_pk=irp_pk)


@gestor_required
def gestao_item_apagar_lote(request, irp_pk):
    irp = get_object_or_404(IRP, pk=irp_pk)
    if irp.fase_atual == 'encerrada':
        messages.error(request, f'A IRP "{irp.titulo}" está encerrada.')
        return redirect('gestao_item_list', irp_pk=irp_pk)
    if request.method == 'POST':
        ids = request.POST.getlist('item_ids')
        if not ids:
            messages.warning(request, 'Nenhum item selecionado.')
        else:
            apagados = Item.objects.filter(pk__in=ids, irp=irp).delete()[0]
            messages.success(request, f'{apagados} item(ns) excluído(s).')
    return redirect('gestao_item_list', irp_pk=irp_pk)


def _norm(s):
    """Normaliza string: maiúsculo, sem acentos, sem espaços extras."""
    if s is None:
        return ''
    s = str(s).strip().upper()
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def _parse_decimal(val):
    """Converte string ou número para Decimal, tolerando vírgula e R$."""
    if val is None:
        return None
    s = str(val).replace('R$', '').replace('\xa0', '').strip()
    if not s:
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _rows_from_xlsx(arquivo):
    """Lê arquivo xlsx e retorna lista de dicts com chaves normalizadas (sem acentos)."""
    import openpyxl
    wb = openpyxl.load_workbook(filename=io.BytesIO(arquivo.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _rows_from_csv(arquivo):
    decoded = arquivo.read().decode('utf-8-sig')
    linhas = decoded.splitlines()
    reader = csv.DictReader(linhas, delimiter=';')
    return [
        {_norm(k): v for k, v in row.items()}
        for row in reader
    ]


def _item_defaults_from_row(row, idx):
    """Extrai campos do item a partir de um dict de linha (chaves normalizadas sem acento)."""
    def get(*keys):
        for k in keys:
            v = row.get(_norm(k))
            if v is not None and str(v).strip():
                return str(v).strip()
        return ''

    numero_raw = get('ITEM', 'NUMERO', 'N')
    try:
        numero = int(float(numero_raw)) if numero_raw else idx
    except (ValueError, TypeError):
        numero = idx

    descricao = get('DISCRIMINACAO DE MATERIAL', 'DESCRICAO', 'DESCRIPTION')
    unidade = get('UND', 'UNIDADE', 'UNIT') or 'UN'
    numero_dfd = get('N DFD', 'NUMERO DFD', 'DFD')
    codigo_catmat = get('CODIGO CATMAT', 'CATMAT', 'COD. CATMAT')

    rubrica_raw = get('RUBRICA', 'NATUREZA', 'TIPO')
    rubrica = rubrica_normalizada(rubrica_raw, strict=bool(rubrica_raw))

    preco_raw = (row.get(_norm('VALOR'))
                 or row.get(_norm('VALOR MAXIMO ACEITAVEL UNITARIO'))
                 or row.get(_norm('PRECO'))
                 or row.get(_norm('PRECO ESTIMADO'))
                 or '0')
    preco = _parse_decimal(preco_raw) or Decimal('0')

    qtd_raw = (row.get(_norm('QUANT. TOTAL'))
               or row.get(_norm('QUANTIDADE TOTAL'))
               or row.get(_norm('QUANTIDADE')))
    quantidade_total = _parse_decimal(qtd_raw)

    return numero, {
        'descricao': descricao,
        'unidade': unidade,
        'rubrica': rubrica,
        'numero_dfd': numero_dfd,
        'codigo_catmat': codigo_catmat,
        'preco_estimado': preco,
        'quantidade_total': quantidade_total,
    }


@gestor_required
def gestao_item_template_xlsx(request, irp_pk):
    """Gera e devolve a planilha modelo para importação de itens."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Itens IRP'

    headers = ['ITEM', 'UND', 'RUBRICA', 'N° DFD', 'CÓDIGO CATMAT', 'DISCRIMINAÇÃO DE MATERIAL', 'VALOR']
    col_widths = [8, 12, 28, 16, 18, 55, 14]

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    rubricas = rubrica_catalog_labels()
    rubrica_exemplo = rubrica_normalizada('Material de Consumo') or (rubricas[0] if rubricas else '')

    # Linha de exemplo
    example = [1, 'UNIDADE', rubrica_exemplo, '774/2024', '233708',
               'PAPEL ALUMÍNIO EM ROLO, LARGURA 45 CM X COMPRIMENTO 7,5 M.', '7,30']
    ex_font = Font(name='Calibri', size=10, italic=True, color='555555')
    for col_idx, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = ex_font
        cell.border = border

    ws.freeze_panes = 'A2'

    if rubricas:
        ws_ref = wb.create_sheet('Rubricas')
        ws_ref['A1'] = 'Rubrica'
        ws_ref['A1'].font = Font(bold=True)
        for row_idx, label in enumerate(rubricas, 2):
            ws_ref.cell(row=row_idx, column=1, value=label)
        ws_ref.column_dimensions['A'].width = 44
        ws_ref.sheet_state = 'hidden'

        dv = DataValidation(type='list', formula1=f"'Rubricas'!$A$2:$A${len(rubricas) + 1}", allow_blank=True)
        dv.errorTitle = 'Rubrica inválida'
        dv.error = 'Escolha uma rubrica cadastrada no módulo Orçamento.'
        dv.promptTitle = 'Rubrica'
        dv.prompt = 'Selecione uma rubrica do cadastro orçamentário.'
        ws.add_data_validation(dv)
        dv.add('C2:C1000')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_importacao_irp.xlsx"'
    wb.save(response)
    return response


@gestor_required
def gestao_item_importar(request, irp_pk):
    irp = get_object_or_404(IRP, pk=irp_pk)
    if irp.fase_atual == 'encerrada':
        messages.error(request, f'A IRP "{irp.titulo}" está encerrada. Não é possível importar itens.')
        return redirect('gestao_item_list', irp_pk=irp_pk)
    if request.method == 'POST':
        form = ItemImportForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = request.FILES['arquivo']
            nome = arquivo.name.lower()
            try:
                if nome.endswith('.xlsx'):
                    linhas = _rows_from_xlsx(arquivo)
                else:
                    linhas = _rows_from_csv(arquivo)

                count = 0
                erros = []
                with transaction.atomic():
                    for i, row in enumerate(linhas, 1):
                        try:
                            numero, defaults = _item_defaults_from_row(row, i)
                            Item.objects.update_or_create(
                                irp=irp, numero=numero,
                                defaults=defaults,
                            )
                            count += 1
                        except Exception as e:
                            erros.append(f'Linha {i}: {e}')

                if erros:
                    messages.warning(request, f'{count} itens importados. Erros: ' + ' | '.join(erros[:5]))
                else:
                    messages.success(request, f'{count} itens importados com sucesso!')
            except Exception as e:
                messages.error(request, f'Erro ao processar arquivo: {e}')
            return redirect('gestao_item_list', irp_pk=irp.pk)
    else:
        form = ItemImportForm()
    return render(request, 'core/gestao/item_importar.html', {
        'form': form,
        'irp': irp,
        'rubricas_validas': rubrica_catalog_labels(),
    })


# ---------------------------------------------------------------------------
# Gestão — Setores
# ---------------------------------------------------------------------------

@gestor_required
def gestao_setor_list(request):
    setores = Setor.objects.all().select_related('pai').order_by('codigo')
    return render(request, 'core/gestao/setor_list.html', {'setores': setores})


@gestor_required
def gestao_setor_create(request):
    if request.method == 'POST':
        form = SetorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Setor criado.')
            return redirect('gestao_setor_list')
    else:
        form = SetorForm()
        pai_id = request.GET.get('pai')
        if pai_id:
            try:
                form.fields['pai'].initial = int(pai_id)
            except (ValueError, TypeError):
                pass
    return render(request, 'core/gestao/setor_form.html', {
        'form': form, 'titulo_pagina': 'Novo Setor'
    })


@gestor_required
def gestao_setor_edit(request, pk):
    setor = get_object_or_404(Setor, pk=pk)
    if request.method == 'POST':
        form = SetorForm(request.POST, instance=setor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Setor atualizado.')
            return redirect('gestao_setor_list')
    else:
        form = SetorForm(instance=setor)
    return render(request, 'core/gestao/setor_form.html', {
        'form': form, 'setor': setor, 'titulo_pagina': 'Editar Setor'
    })


@gestor_required
def gestao_setor_delete(request, pk):
    """Desativa um setor (soft delete)."""
    setor = get_object_or_404(Setor, pk=pk)
    if request.method == 'POST':
        setor.ativo = False
        setor.save()
        messages.success(request, f'Setor "{setor.nome}" desativado.')
    return redirect('gestao_setor_list')


def _setor_tem_dependencias(setor):
    """Retorna mensagem de erro se o setor tiver dependências, ou None."""
    if setor.subsetores.exists():
        return f'"{setor.nome}" possui subsetores vinculados'
    if PerfilUsuario.objects.filter(setor=setor).exists():
        return f'"{setor.nome}" está vinculado a usuários'
    if Resposta.objects.filter(setor=setor).exists():
        return f'"{setor.nome}" está vinculado a respostas de IRP'
    return None


@gestor_required
def gestao_setor_apagar(request, pk):
    """Apaga permanentemente um setor (verifica dependências antes)."""
    setor = get_object_or_404(Setor, pk=pk)
    if request.method == 'POST':
        erro = _setor_tem_dependencias(setor)
        if erro:
            messages.error(request, f'Não é possível apagar: {erro}.')
        else:
            nome = setor.nome
            setor.delete()
            messages.success(request, f'Setor "{nome}" apagado permanentemente.')
    return redirect('gestao_setor_list')


@gestor_required
def gestao_setor_apagar_lote(request):
    """Apaga permanentemente os setores selecionados."""
    if request.method == 'POST':
        ids = request.POST.getlist('setor_ids')
        if not ids:
            messages.warning(request, 'Nenhum setor selecionado.')
            return redirect('gestao_setor_list')
        apagados, erros = 0, []
        for pk in ids:
            try:
                setor = Setor.objects.get(pk=pk)
                erro = _setor_tem_dependencias(setor)
                if erro:
                    erros.append(erro)
                else:
                    setor.delete()
                    apagados += 1
            except Setor.DoesNotExist:
                pass
        if apagados:
            messages.success(request, f'{apagados} setor(es) apagado(s) permanentemente.')
        for erro in erros:
            messages.error(request, f'Nao foi possivel apagar: {erro}.')
    return redirect('gestao_setor_list')


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Recuperação de usuário / senha
# ---------------------------------------------------------------------------

def recuperar_usuario(request):
    mensagem = None
    erro = None
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if email:
            users = User.objects.filter(email__iexact=email)
            if users.exists():
                for u in users:
                    try:
                        send_mail(
                            subject='Recuperação de usuário – IRP CT/UFPB',
                            message=f'Olá,\n\nSeu usuário de acesso ao Sistema IRP CT/UFPB é: {u.username}\n\nCaso não tenha solicitado esta recuperação, ignore este e-mail.\n\nAtenciosamente,\nSistema IRP CT/UFPB',
                            from_email=None,
                            recipient_list=[email],
                            fail_silently=False,
                        )
                    except Exception:
                        pass
            # Always show success (security: don't reveal if email exists)
            mensagem = 'Se o endereço informado estiver cadastrado, você receberá um e-mail com seu usuário em breve.'
        else:
            erro = 'Informe um endereço de e-mail.'
    return render(request, 'registration/recuperar_usuario.html', {'mensagem': mensagem, 'erro': erro})


def recuperar_senha(request):
    mensagem = None
    erro = None
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if email:
            users = User.objects.filter(email__iexact=email)
            if users.exists():
                for u in users:
                    try:
                        uid = urlsafe_base64_encode(force_bytes(u.pk))
                        token = default_token_generator.make_token(u)
                        link = request.build_absolute_uri(
                            f'/recuperar-senha/confirmar/{uid}/{token}/'
                        )
                        send_mail(
                            subject='Redefinição de senha – IRP CT/UFPB',
                            message=f'Olá, {u.username},\n\nClique no link abaixo para redefinir sua senha (válido por 30 minutos):\n\n{link}\n\nCaso não tenha solicitado, ignore este e-mail.\n\nAtenciosamente,\nSistema IRP CT/UFPB',
                            from_email=None,
                            recipient_list=[email],
                            fail_silently=False,
                        )
                    except Exception:
                        pass
            mensagem = 'Se o endereço informado estiver cadastrado, você receberá um link para redefinição de senha em breve.'
        else:
            erro = 'Informe um endereço de e-mail.'
    return render(request, 'registration/recuperar_senha.html', {'mensagem': mensagem, 'erro': erro})


def confirmar_senha(request, uidb64, token):
    erro = None
    sucesso = False
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is None or not default_token_generator.check_token(user, token):
        erro = 'Este link é inválido ou expirou. Solicite uma nova redefinição de senha.'
        return render(request, 'registration/confirmar_senha.html', {'erro': erro, 'token_valido': False})

    if request.method == 'POST':
        p1 = request.POST.get('password1', '')
        p2 = request.POST.get('password2', '')
        if not p1:
            erro = 'Informe a nova senha.'
        elif p1 != p2:
            erro = 'As senhas não coincidem.'
        elif len(p1) < 8:
            erro = 'A senha deve ter pelo menos 8 caracteres.'
        else:
            user.set_password(p1)
            user.save()
            sucesso = True
    return render(request, 'registration/confirmar_senha.html', {
        'erro': erro, 'sucesso': sucesso, 'token_valido': True,
        'uidb64': uidb64, 'token': token,
    })


def ativar_conta(request, uidb64, token):
    """
    View para ativação de nova conta (define a primeira senha e ativa is_active=True).
    Reutiliza a lógica de confirmar_senha mas com contexto de boas-vindas.
    """
    erro = None
    sucesso = False
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is None or not default_token_generator.check_token(user, token):
        erro = 'Este link de ativação é inválido ou expirou. Entre em contato com o administrador.'
        return render(request, 'registration/ativar_conta.html', {'erro': erro, 'token_valido': False})
    
    # Se o usuário já estiver ativo, talvez já tenha ativado?
    if user.is_active:
        messages.info(request, "Sua conta já está ativa. Se esqueceu a senha, use a recuperação de senha.")
        return redirect('login')

    if request.method == 'POST':
        p1 = request.POST.get('password1', '')
        p2 = request.POST.get('password2', '')
        if not p1:
            erro = 'Informe uma senha.'
        elif p1 != p2:
            erro = 'As senhas não coincidem.'
        elif len(p1) < 8:
            erro = 'A senha deve ter pelo menos 8 caracteres.'
        else:
            user.set_password(p1)
            user.is_active = True # Ativa a conta aqui!
            user.save()
            sucesso = True
            messages.success(request, "Conta ativada com sucesso! Você já pode entrar no sistema.")

    return render(request, 'registration/ativar_conta.html', {
        'erro': erro, 'sucesso': sucesso, 'token_valido': True,
        'uidb64': uidb64, 'token': token,
        'username': user.username
    })


# Gestão — Usuários
# ---------------------------------------------------------------------------

@gestor_required
def gestao_usuario_list(request):
    usuarios = User.objects.select_related('perfil').order_by('perfil__nome_completo', 'username')
    return render(request, 'core/gestao/usuario_list.html', {'usuarios': usuarios})


@gestor_required
def gestao_usuario_create(request):
    todos_setores = Setor.objects.filter(ativo=True).order_by('nome')
    if request.method == 'POST':
        form = UsuarioForm(request.POST, request.FILES)
        if form.is_valid():
            with transaction.atomic():
                user = form.save()
                
                # Gerar link de ativação
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                token = default_token_generator.make_token(user)
                link = request.build_absolute_uri(f'/ativar-conta/{uid}/{token}/')
                
                # Enviar e-mail
                assunto = 'Ative sua conta – IRP CT/UFPB'
                mensagem = (
                    f'Olá, {user.username},\n\n'
                    'Sua conta no Sistema IRP CT/UFPB foi criada com sucesso.\n'
                    'Para começar a utilizar o sistema, você precisa ativar sua conta e definir sua senha '
                    'clicando no link abaixo (válido por 24 horas):\n\n'
                    f'{link}\n\n'
                    'Após definir sua senha, você poderá acessar o sistema com seu usuário e a senha cadastrada.\n\n'
                    'Atenciosamente,\nSistema IRP CT/UFPB'
                )
                try:
                    send_mail(
                        subject=assunto,
                        message=mensagem,
                        from_email=None,
                        recipient_list=[user.email],
                        fail_silently=False,
                    )
                    messages.success(request, f'Usuário "{user.username}" criado. Um e-mail de ativação foi enviado para {user.email}.')
                except Exception as e:
                    messages.warning(request, f'Usuário "{user.username}" criado, mas houve um erro ao enviar o e-mail de ativação: {e}')
                
            return redirect('gestao_usuario_list')
    else:
        form = UsuarioForm()
    return render(request, 'core/gestao/usuario_form.html', {
        'form': form, 'titulo_pagina': 'Novo Usuário',
        'todos_setores': todos_setores,
        'tipo_setor_choices': TIPO_SETOR_CHOICES,
        'perfil_por_tipo_json': json.dumps(PERFIL_POR_TIPO_SETOR),
    })


@gestor_required
def gestao_usuario_edit(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    try:
        perfil = usuario.perfil
    except PerfilUsuario.DoesNotExist:
        perfil = None

    todos_setores = Setor.objects.filter(ativo=True).order_by('nome')

    if request.method == 'POST':
        form = UsuarioForm(request.POST, request.FILES, instance=usuario, perfil=perfil)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário atualizado.')
            return redirect('gestao_usuario_list')
    else:
        form = UsuarioForm(instance=usuario, perfil=perfil)

    return render(request, 'core/gestao/usuario_form.html', {
        'form': form, 'usuario': usuario, 'titulo_pagina': 'Editar Usuário',
        'todos_setores': todos_setores,
        'tipo_setor_choices': TIPO_SETOR_CHOICES,
        'perfil': perfil,
        'perfil_por_tipo_json': json.dumps(PERFIL_POR_TIPO_SETOR),
    })


@gestor_required
def gestao_usuario_toggle_ativo(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST' and usuario != request.user:
        usuario.is_active = not usuario.is_active
        usuario.save()
        estado = 'ativado' if usuario.is_active else 'desativado'
        messages.success(request, f'Usuário {estado}.')
    return redirect('gestao_usuario_list')


@gestor_required
def gestao_usuario_apagar(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        if usuario == request.user:
            messages.error(request, 'Você não pode apagar sua própria conta.')
            return redirect('gestao_usuario_list')
        nome = getattr(getattr(usuario, 'perfil', None), 'nome_completo', None) or usuario.username
        usuario.delete()
        messages.success(request, f'Usuário "{nome}" apagado permanentemente.')
    return redirect('gestao_usuario_list')


@gestor_required
def gestao_usuario_apagar_lote(request):
    if request.method == 'POST':
        ids = request.POST.getlist('usuario_ids')
        apagados = 0
        for uid in ids:
            try:
                u = User.objects.get(pk=int(uid))
                if u != request.user:
                    u.delete()
                    apagados += 1
            except (User.DoesNotExist, ValueError):
                pass
        if apagados:
            messages.success(request, f'{apagados} usuário(s) apagado(s) permanentemente.')
    return redirect('gestao_usuario_list')


# ---------------------------------------------------------------------------
# Gestão — Resultados e Exportação
# ---------------------------------------------------------------------------

@gestor_required
def gestao_resultados(request, irp_pk):
    irp = get_object_or_404(IRP, pk=irp_pk)
    respostas = (
        Resposta.objects
        .filter(irp=irp)
        .select_related('usuario__perfil', 'setor')
        .prefetch_related('itens_resposta__item')
        .order_by('setor__nome', 'usuario__perfil__nome_completo')
    )

    # Enriquece cada resposta com totais
    respostas_dados = []
    for resp in respostas:
        total_v = Decimal('0')
        total_i = 0
        for ri in resp.itens_resposta.all():
            if ri.quantidade and ri.quantidade > 0:
                total_v += ri.quantidade * ri.item.preco_estimado
                total_i += 1
        respostas_dados.append({
            'resposta': resp,
            'total_valor': total_v,
            'total_itens': total_i,
        })

    dados = _calcular_dados([irp])

    # Summary stats derived directly from respostas_dados for card accuracy
    total_respondentes = len(respostas_dados)
    total_itens_intencionados = sum(rd['total_itens'] for rd in respostas_dados)
    total_valor = sum(rd['total_valor'] for rd in respostas_dados)

    # Chart data: aggregate by setor, grouped by level
    por_nivel_chart = _build_por_nivel_chart(
        resp['resposta'] for resp in respostas_dados
    )

    return render(request, 'core/gestao/resultados.html', {
        'irp': irp,
        'respostas_dados': respostas_dados,
        'dados': dados,
        'total_respondentes': total_respondentes,
        'total_itens_intencionados': total_itens_intencionados,
        'total_valor': total_valor,
        'por_nivel_chart': por_nivel_chart,
    })


@gestor_required
def gestao_exportar(request, irp_pk):
    irp = get_object_or_404(IRP, pk=irp_pk)

    # Agregação de dados por item (Intencionado e Homologado)
    from django.db.models import Q, Sum
    items_agg = irp.itens.annotate(
        q_intenc=Sum('respostas_item__quantidade', filter=Q(respostas_item__resposta__respondida_em__isnull=False)),
        q_hom=Sum('homologacoes_setor__quantidade_aprovada', filter=Q(homologacoes_setor__homologacao__status='homologada'))
    ).order_by('numero')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Estilos comuns
    header_font    = Font(bold=True, color='FFFFFF', size=10)
    header_fill    = PatternFill('solid', fgColor='1D4ED8')
    center         = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left           = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin           = Side(style='thin', color='CCCCCC')
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Aba 1: Síntese para a PRA ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Síntese para a PRA"
    
    headers1 = ['ITEM', 'UND', 'RUBRICA', 'N° DFD', 'CÓDIGO CATMAT', 'DISCRIMINAÇÃO DE MATERIAL', 'VALOR', 'QTE INTENCIONADA', 'QTE HOMOLOGADA']
    ws1.append(headers1)
    for col, _ in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws1.row_dimensions[1].height = 30

    row_num1 = 2
    for item in items_agg:
        rubrica = item.get_rubrica_display() if item.rubrica else ''
        linha1 = [
            item.numero,
            item.unidade,
            rubrica,
            item.numero_dfd,
            item.codigo_catmat,
            item.descricao,
            float(item.preco_estimado),
            float(item.q_intenc or 0),
            float(item.q_hom or 0),
        ]
        ws1.append(linha1)
        for col, val in enumerate(linha1, 1):
            cell = ws1.cell(row=row_num1, column=col)
            cell.border = border
            cell.alignment = center if col in (1, 2, 4, 5, 8, 9) else left
            if col in (7, 8, 9):
                cell.number_format = '#,##0.00'
        row_num1 += 1

    larguras1 = [8, 10, 25, 15, 15, 60, 14, 18, 18]
    for i, w in enumerate(larguras1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'

    # ── Aba 2: Respostas do CT ─────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Respostas do CT")
    cabecalho2 = [
        'IRP', 'Setor',
        'Nome do Respondente', 'Matrícula',
        'Nº Item', 'Rubrica', 'Descrição do Item', 'Unidade',
        'Preço Estimado (R$)', 'Quantidade Intencionada',
        'Valor Total (R$)', 'Observação', 'Última Atualização',
    ]
    ws2.append(cabecalho2)
    for col, _ in enumerate(cabecalho2, 1):
        cell = ws2.cell(row=1, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws2.row_dimensions[1].height = 30

    respostas = (
        Resposta.objects
        .filter(irp=irp, respondida_em__isnull=False)
        .select_related('usuario__perfil', 'setor')
        .prefetch_related('itens_resposta__item')
        .order_by('setor__nome')
    )

    row_num2 = 2
    for resp in respostas:
        try:
            nome      = resp.usuario.perfil.nome_completo
            matricula = resp.usuario.perfil.matricula or ''
        except Exception:
            nome      = resp.usuario.get_full_name() or resp.usuario.username
            matricula = ''

        for ri in resp.itens_resposta.filter(quantidade__gt=0).order_by('item__numero'):
            valor = ri.quantidade * ri.item.preco_estimado
            linha2 = [
                irp.titulo,
                resp.setor.nome if resp.setor else '',
                nome, matricula,
                ri.item.numero,
                ri.item.get_rubrica_display() if ri.item.rubrica else '',
                ri.item.descricao,
                ri.item.unidade,
                float(ri.item.preco_estimado or 0),
                float(ri.quantidade or 0),
                float(valor or 0),
                ri.observacao or '',
                resp.atualizada_em.strftime('%d/%m/%Y %H:%M'),
            ]
            ws2.append(linha2)
            # Aplicar estilos à linha recém adicionada (ws2.max_row)
            m_row = ws2.max_row
            for col_idx, _ in enumerate(linha2, 1):
                cell = ws2.cell(row=m_row, column=col_idx)
                cell.border = border
                cell.alignment = center if col_idx in (1, 5, 8, 9, 10, 11) else left
                if col_idx in (9, 10, 11):
                    cell.number_format = '#,##0.00'

    larguras2 = [30, 30, 30, 15, 8, 25, 60, 10, 16, 18, 16, 35, 18]
    for i, w in enumerate(larguras2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    # ── Finalização ────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"IRP_{irp.titulo.replace('/', '.')}_respostas_homologadas.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
